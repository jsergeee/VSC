# school/admin.py

from django.contrib import admin, messages
from django.core.exceptions import ValidationError
from django.contrib.auth.admin import UserAdmin
from django.utils.html import format_html
from django.contrib.auth.models import User as AuthUser
from django.urls import path
from django.shortcuts import render, redirect
from django import forms
from django.db import models
from django.utils import timezone
from datetime import datetime
from django.db.models import Prefetch, Sum, Count
from django.db import transaction
from .views import import_users_view
from .telegram import notify_payment
# В начале файла оставьте ТОЛЬКО ОДИН импорт всех моделей
from .models import (
    User, Subject, Teacher, Student, Lesson, LessonFormat,
    LessonReport, Payment, Schedule, TrialRequest,
    Notification, LessonFeedback, TeacherRating,
    Homework, HomeworkSubmission, GroupLesson, GroupEnrollment,
    LessonAttendance, ScheduleTemplate, ScheduleTemplateStudent,
    StudentSubjectPrice, UserActionLog, PaymentRequest, Feedback
)
from .views import schedule_calendar_data, admin_complete_lesson

# Импортируем helper-классы для финансовых расчетов
from .views import LessonFinanceCalculator, PeriodFinanceCalculator

# Разрегистрируем стандартного User, если зарегистрирован
try:
    admin.site.unregister(AuthUser)
except admin.sites.NotRegistered:
    pass

"""
Модуль администрирования для приложения school.

Содержит настройки отображения всех моделей в Django admin панели,
кастомные действия, инлайн-формы и методы обработки.

Основные компоненты:
    * Кастомный UserAdmin — расширенное управление пользователями
    * TeacherAdmin — управление учителями с финансовой статистикой
    * StudentAdmin — управление учениками с балансом и отчетностью
    * LessonAdmin — управление уроками с проверкой занятости учителя
    * PaymentAdmin — управление платежами с автоматической корректировкой баланса
    * PaymentRequestAdmin — управление запросами на выплаты учителям
    * HomeworkAdmin — управление домашними заданиями
    * И другие ModelAdmin классы для всех моделей проекта

Также содержит кастомные действия (actions), инлайн-формы,
методы для отображения статистики и интеграцию с Telegram уведомлениями.
"""
# ==================== INLINES ====================

class StudentSubjectPriceInline(admin.TabularInline):
    model = StudentSubjectPrice
    extra = 1
    fields = ['teacher', 'subject', 'cost', 'teacher_payment', 'discount', 'is_active']
    autocomplete_fields = ['subject', 'teacher']


class LessonAttendanceInline(admin.TabularInline):
    model = LessonAttendance
    extra = 1
    raw_id_fields = ['student']
    fields = ['student', 'cost', 'discount', 'teacher_payment_share', 'status']


class GroupEnrollmentInline(admin.TabularInline):
    model = GroupEnrollment
    extra = 1
    raw_id_fields = ['student']


class ScheduleTemplateStudentInline(admin.TabularInline):
    model = ScheduleTemplateStudent
    extra = 1
    raw_id_fields = ['student']


# ==================== CUSTOM USER ADMIN ====================

class CustomUserAdmin(UserAdmin):
    """
    Административный интерфейс для модели User.

    Расширяет стандартный UserAdmin Django, добавляя поля для ролей,
    подтверждения email, Telegram уведомлений и кастомные действия.

    Атрибуты:
        list_display: Поля, отображаемые в списке пользователей
        list_filter: Поля для фильтрации
        search_fields: Поля для поиска
        fieldsets: Группировка полей в форме редактирования
        add_fieldsets: Группировка полей в форме добавления
        actions: Доступные массовые действия

    Кастомные методы:
        enable_telegram_notifications() - включает Telegram уведомления
        disable_telegram_notifications() - отключает Telegram уведомления
        mark_as_verified() - отмечает email как подтвержденный
        mark_as_unverified() - отмечает email как неподтвержденный
        export_users_excel() - экспортирует пользователей в Excel
        get_full_name() - возвращает полное имя пользователя
        is_email_verified_badge() - отображает статус email в виде цветного значка

    Кастомные действия (actions):
        * Включение/отключение Telegram уведомлений
        * Подтверждение/снятие подтверждения email
        * Экспорт в Excel
    """
    list_display = ('id', 'username', 'get_full_name', 'email', 'phone', 'role',
                    'is_email_verified_badge', 'is_staff', 'telegram_notifications')
    list_filter = ('telegram_notifications', 'role', 'is_email_verified', 'is_staff', 'is_superuser', 'groups')
    search_fields = ('username', 'last_name', 'first_name',  'email', 'phone', 'telegram_chat_id')
    readonly_fields = ('email_verification_sent',)

    fieldsets = (
        (None, {'fields': ('username', 'password')}),
        ('Личная информация', {
            'fields': ('last_name', 'first_name',  'patronymic', 'email', 'phone', 'photo')
        }),
        ('Роль', {
            'fields': ('role',),
            'classes': ('wide',),
        }),
        ('✅ Email подтверждение', {
            'fields': ('is_email_verified', 'email_verification_sent'),
            'classes': ('wide',),
            'description': 'Управление подтверждением email пользователя',
        }),
        ('Права доступа', {
            'fields': ('is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions'),
        }),
        ('Важные даты', {'fields': ('last_login', 'date_joined')}),
        ('Telegram уведомления', {
            'fields': (
                'telegram_chat_id',
                'telegram_notifications',
            ),
            'classes': ('wide',),
            'description': 'Настройки уведомлений в Telegram'
        }),
    )

    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': ('username', 'password1', 'password2',
                       'first_name', 'last_name', 'patronymic',
                       'email', 'phone', 'role'),
        }),
        ('Telegram уведомления', {
            'fields': (
                'telegram_chat_id',
                'telegram_notifications',
            ),
        }),
    )

    actions = ['mark_as_verified', 'mark_as_unverified', 'export_users_excel', 'import_users_excel', 'enable_telegram_notifications', 'disable_telegram_notifications']

    def enable_telegram_notifications(self, request, queryset):
        updated = queryset.update(telegram_notifications=True)
        self.message_user(request, f'Уведомления включены для {updated} пользователей')

    enable_telegram_notifications.short_description = "Включить Telegram уведомления"

    def disable_telegram_notifications(self, request, queryset):
        updated = queryset.update(telegram_notifications=False)
        self.message_user(request, f'Уведомления отключены для {updated} пользователей')

    disable_telegram_notifications.short_description = "Отключить Telegram уведомления"

    def get_urls(self):
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('import/', self.admin_site.admin_view(import_users_view), name='import_users'),
        ]
        return custom_urls + urls

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['show_import_button'] = True
        return super().changelist_view(request, extra_context)

    def get_full_name(self, obj):
        """Возвращает полное имя пользователя (использует метод модели)"""
        return obj.get_full_name() or obj.username

    get_full_name.short_description = 'ФИО'

    def is_email_verified_badge(self, obj):
        if obj.is_email_verified:
            return format_html(
                '<span style="background: #28a745; color: white; padding: 3px 8px; border-radius: 3px;">✅ Подтвержден</span>'
            )
        else:
            return format_html(
                '<span style="background: #dc3545; color: white; padding: 3px 8px; border-radius: 3px;">❌ Не подтвержден</span>'
            )

    is_email_verified_badge.short_description = 'Email подтвержден'

    def mark_as_verified(self, request, queryset):
        updated = queryset.update(is_email_verified=True)
        self.message_user(request, f'✅ {updated} пользователей отмечены как подтвержденные')

    mark_as_verified.short_description = "✅ Отметить как подтвержденные email"

    def mark_as_unverified(self, request, queryset):
        updated = queryset.update(is_email_verified=False)
        self.message_user(request, f'⚠️ {updated} пользователей отмечены как неподтвержденные')

    mark_as_unverified.short_description = "❌ Отметить как неподтвержденные email"

    def export_users_excel(self, request, queryset):
        """Экспорт пользователей в Excel"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        from datetime import datetime

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Пользователи"

        headers = ['ID', 'Логин', 'Фамилия', 'Имя', 'Отчество', 'Email', 'Телефон', 'Роль']

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row, user in enumerate(queryset, start=2):
            ws.cell(row=row, column=1, value=user.id)
            ws.cell(row=row, column=2, value=user.username)
            ws.cell(row=row, column=3, value=user.last_name)
            ws.cell(row=row, column=4, value=user.first_name)
            ws.cell(row=row, column=5, value=user.patronymic)
            ws.cell(row=row, column=6, value=user.email)
            ws.cell(row=row, column=7, value=user.phone)
            ws.cell(row=row, column=8, value=user.get_role_display())

        column_widths = [8, 15, 15, 15, 15, 25, 15, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"users_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    export_users_excel.short_description = "📥 Экспорт выбранных пользователей в Excel"


# ==================== SUBJECT ADMIN ====================

class SubjectAdmin(admin.ModelAdmin):
    list_display = ('name', 'description', 'teachers_count', 'lessons_count')
    search_fields = ('name',)

    def teachers_count(self, obj):
        return obj.teacher_set.count()

    teachers_count.short_description = 'Учителей'

    def lessons_count(self, obj):
        return obj.lesson_set.count()

    lessons_count.short_description = 'Занятий'


# ==================== TEACHER ADMIN ====================

    @admin.register(Teacher)
    class TeacherAdmin(admin.ModelAdmin):
        list_display = (
            'id',
            'user_link',
            'user_photo_preview',
            'display_subjects',
            'experience',
            'students_count',
            'rating_display',
            'get_telegram_status',
            'show_on_team',
            'team_sort_order',
            'team_highlight',
            'show_on_team_badge',
            'team_highlight_badge'
        )
        list_filter = ('subjects', 'show_on_team', 'team_highlight')
        list_editable = ('show_on_team', 'team_sort_order', 'team_highlight')
        search_fields = ('user__first_name', 'user__last_name', 'user__email')
        filter_horizontal = ('subjects',)
        readonly_fields = ('wallet_balance', 'rating_display')
        change_list_template = "admin/school/teacher/change_list_with_period.html"

        fieldsets = (
            (None, {
                'fields': ('user', 'subjects', 'experience')
            }),
            ('Финансы', {
                'fields': ('wallet_balance', 'payment_details'),
                'classes': ('wide',),
            }),
            ('Дополнительная информация', {
                'fields': ('education', 'bio', 'certificate'),
                'classes': ('collapse',),
            }),
            ('Рейтинг', {
                'fields': ('rating_display',),
                'classes': ('wide',),
            }),
            ('👥 Страница команды', {
                'fields': (
                    'show_on_team',
                    'team_sort_order',
                    'team_highlight',
                    'team_description',
                    # 'team_photo', ← если есть, удалите или закомментируйте
                ),
                'classes': ('wide', 'collapse'),
                'description': 'Настройки отображения на странице /team/',
            }),
            ('🌐 Социальные сети', {
                'fields': (
                    'telegram_url',
                    'whatsapp_url',
                    'vk_url',
                    'instagram_url',
                ),
                'classes': ('wide', 'collapse'),
            }),
        )

        def user_photo_preview(self, obj):
            if obj.user.photo:
                return format_html(
                    '<img src="{}" style="max-height: 50px; max-width: 50px; border-radius: 5px; object-fit: cover;" />',
                    obj.user.photo.url
                )
            return format_html(
                '<div style="width: 50px; height: 50px; background: #f0f0f0; border-radius: 5px; display: flex; align-items: center; justify-content: center;">'
                '<i class="fa fa-user" style="color: #999;"></i>'
                '</div>'
            )

        user_photo_preview.short_description = 'Фото'
        def get_telegram_status(self, obj):
            if obj.user.telegram_chat_id:
                return f"✅ {obj.user.telegram_chat_id}"
            return "❌ Не подключен"

        get_telegram_status.short_description = "Telegram"

        def user_link(self, obj):
            url = f'/admin/school/user/{obj.user.id}/change/'
            return format_html('<a href="{}">{}</a>', url, obj.user.get_full_name())

        user_link.short_description = 'Учитель'

        def display_subjects(self, obj):
            return ", ".join([s.name for s in obj.subjects.all()])

        display_subjects.short_description = 'Предметы'

        def students_count(self, obj):
            return obj.student_set.count()

        students_count.short_description = 'Учеников'

        def rating_display(self, obj):
            try:
                rating = obj.rating_stats
                stars = '⭐' * int(rating.average_rating)
                return f"{stars} ({rating.average_rating:.1f}) - {rating.total_feedbacks} оценок"
            except:
                return 'Нет оценок'

        rating_display.short_description = 'Рейтинг'

        def show_on_team_badge(self, obj):
            if obj.show_on_team:
                return format_html(
                    '<span style="background: #28a745; color: white; padding: 3px 8px; border-radius: 3px;">✅ На сайте</span>'
                )
            return format_html(
                '<span style="background: #dc3545; color: white; padding: 3px 8px; border-radius: 3px;">❌ Скрыт</span>'
            )

        show_on_team_badge.short_description = 'На сайте'

        def team_highlight_badge(self, obj):
            if obj.team_highlight:
                return format_html(
                    '<span style="background: #ffc107; color: #333; padding: 3px 8px; border-radius: 3px;">⭐ Выделен</span>'
                )
            return ''

        team_highlight_badge.short_description = 'Выделен'

        # ⚡⚡⚡ СОХРАНЯЕМ REQUEST ДЛЯ ИСПОЛЬЗОВАНИЯ В ДРУГИХ МЕТОДАХ ⚡⚡⚡
        def get_queryset(self, request):
            self.request = request
            return super().get_queryset(request)

        # ⚡⚡⚡ МЕТОД ДЛЯ ОБРАБОТКИ СПИСКА (ТОЛЬКО ДЛЯ ТАБЛИЦЫ СТАТИСТИКИ) ⚡⚡⚡
        def changelist_view(self, request, extra_context=None):
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')

            # Если параметры есть в GET, сохраняем их в сессию
            if start_date and end_date:
                request.session['teacher_filter_start'] = start_date
                request.session['teacher_filter_end'] = end_date
            else:
                # Если параметров нет, пробуем взять из сессии
                start_date = request.session.get('teacher_filter_start')
                end_date = request.session.get('teacher_filter_end')

            print("\n" + "=" * 80)
            print("🔍 TEACHER ADMIN CHANGELIST VIEW")
            print(f"📅 start_date: {start_date}")
            print(f"📅 end_date: {end_date}")
            print("=" * 80)

            if start_date and end_date:
                try:
                    from datetime import datetime
                    start = datetime.strptime(start_date, '%Y-%m-%d').date()
                    end = datetime.strptime(end_date, '%Y-%m-%d').date()

                    print(f"\n✅ Период преобразован: {start} - {end}")

                    extra_context = extra_context or {}
                    teachers_data = []

                    # Получаем всех учителей
                    teachers = self.get_queryset(request)
                    print(f"\n👥 Всего учителей: {teachers.count()}")

                    for teacher in teachers:
                        print(f"\n{'─' * 50}")
                        print(f"👨‍🏫 Обработка учителя: {teacher.user.get_full_name()} (ID: {teacher.id})")

                        # Получаем статистику
                        earnings = teacher.get_teacher_earnings(start, end)

                        print(f"📊 earnings получены:")
                        print(f"   total_payments: {earnings.get('total_payments', 0)}")
                        print(f"   total_salaries: {earnings.get('total_salaries', 0)}")
                        print(f"   net_income: {earnings.get('net_income', 0)}")
                        print(f"   payments_count: {earnings.get('payments_count', 0)}")
                        print(f"   salaries_count: {earnings.get('salaries_count', 0)}")

                        teachers_data.append({
                            'teacher': teacher,
                            'earnings': earnings
                        })

                    extra_context['teachers_data'] = teachers_data
                    extra_context['start_date'] = start_date
                    extra_context['end_date'] = end_date

                    print(f"\n✅ teachers_data создан, размер: {len(teachers_data)}")

                    # Выводим итоговые данные для всех учителей
                    print(f"\n📊 ИТОГОВЫЕ ДАННЫЕ:")
                    for i, item in enumerate(teachers_data):
                        teacher_name = item['teacher'].user.get_full_name()
                        earnings = item['earnings']
                        print(f"{i + 1}. {teacher_name}:")
                        print(f"   💰 total_payments: {earnings.get('total_payments', 0)}")
                        print(f"   💵 total_salaries: {earnings.get('total_salaries', 0)}")
                        print(f"   💹 net_income: {earnings.get('net_income', 0)}")

                except Exception as e:
                    print(f"❌ ОШИБКА в changelist_view: {e}")
                    import traceback
                    traceback.print_exc()

            print("=" * 80 + "\n")
            return super().changelist_view(request, extra_context)

        actions = ['export_teachers_excel', 'calculate_payments',
                   'show_on_team', 'hide_from_team', 'set_highlight', 'remove_highlight']

        def export_teachers_excel(self, request, queryset):
            """Экспорт выбранных учителей в Excel"""
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from django.http import HttpResponse
            from datetime import datetime
            from school.models import Lesson

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Учителя"

            headers = ['ID', 'Фамилия', 'Имя', 'Отчество', 'Email', 'Телефон',
                       'Предметы', 'Опыт', 'Всего уроков', 'Проведено уроков', 'Учеников']

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            for row, teacher in enumerate(queryset, start=2):
                subjects = ", ".join([s.name for s in teacher.subjects.all()])

                total_lessons = Lesson.objects.filter(teacher=teacher).count()
                completed_lessons = Lesson.objects.filter(teacher=teacher, status='completed').count()
                total_students = teacher.student_set.count()

                ws.cell(row=row, column=1, value=teacher.id)
                ws.cell(row=row, column=2, value=teacher.user.last_name)
                ws.cell(row=row, column=3, value=teacher.user.first_name)
                ws.cell(row=row, column=4, value=teacher.user.patronymic)
                ws.cell(row=row, column=5, value=teacher.user.email)
                ws.cell(row=row, column=6, value=teacher.user.phone)
                ws.cell(row=row, column=7, value=subjects)
                ws.cell(row=row, column=8, value=teacher.experience)
                ws.cell(row=row, column=9, value=total_lessons)
                ws.cell(row=row, column=10, value=completed_lessons)
                ws.cell(row=row, column=11, value=total_students)

            column_widths = [8, 15, 15, 15, 25, 15, 30, 8, 12, 15, 10]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"teachers_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            wb.save(response)
            return response

        export_teachers_excel.short_description = "📥 Экспорт выбранных учителей в Excel"

        def calculate_payments(self, request, queryset):
            """Перейти к расчету выплат"""
            if queryset.count() == 1:
                teacher = queryset.first()
                return redirect(f'/admin/teacher-payments/?teacher_id={teacher.id}')
            else:
                self.message_user(request, 'Выберите одного учителя для расчета выплат', level='WARNING')

        calculate_payments.short_description = "💰 Расчет выплат"

        # ACTIONS ДЛЯ СТРАНИЦЫ КОМАНДЫ
        def show_on_team(self, request, queryset):
            updated = queryset.update(show_on_team=True)
            self.message_user(request, f'✅ {updated} учителей будут показываться на странице команды')

        show_on_team.short_description = "✅ Показывать на странице команды"

        def hide_from_team(self, request, queryset):
            updated = queryset.update(show_on_team=False)
            self.message_user(request, f'❌ {updated} учителей скрыты со страницы команды')

        hide_from_team.short_description = "❌ Скрыть со страницы команды"

        def set_highlight(self, request, queryset):
            updated = queryset.update(team_highlight=True)
            self.message_user(request, f'⭐ {updated} учителей отмечены как "выделенные"')

        set_highlight.short_description = "⭐ Отметить как выделенных"

        def remove_highlight(self, request, queryset):
            updated = queryset.update(team_highlight=False)
            self.message_user(request, f'➖ У {updated} учителей снята отметка "выделенные"')

        remove_highlight.short_description = "➖ Снять выделение"


    def get_telegram_status(self, obj):
        if obj.user.telegram_chat_id:
            return f"✅ {obj.user.telegram_chat_id}"
        return "❌ Не подключен"
    get_telegram_status.short_description = "Telegram"

    def user_link(self, obj):
        url = f'/admin/school/user/{obj.user.id}/change/'
        return format_html('<a href="{}">{}</a>', url, obj.user.get_full_name())
    user_link.short_description = 'Учитель'

    def display_subjects(self, obj):
        return ", ".join([s.name for s in obj.subjects.all()])
    display_subjects.short_description = 'Предметы'

    def students_count(self, obj):
        return obj.student_set.count()
    students_count.short_description = 'Учеников'

    def rating_display(self, obj):
        try:
            rating = obj.rating_stats
            stars = '⭐' * int(rating.average_rating)
            return f"{stars} ({rating.average_rating:.1f}) - {rating.total_feedbacks} оценок"
        except:
            return 'Нет оценок'
    rating_display.short_description = 'Рейтинг'

    def show_on_team_badge(self, obj):
        if obj.show_on_team:
            return format_html(
                '<span style="background: #28a745; color: white; padding: 3px 8px; border-radius: 3px;">✅ На сайте</span>'
            )
        return format_html(
            '<span style="background: #dc3545; color: white; padding: 3px 8px; border-radius: 3px;">❌ Скрыт</span>'
        )
    show_on_team_badge.short_description = 'На сайте'

    def team_highlight_badge(self, obj):
        if obj.team_highlight:
            return format_html(
                '<span style="background: #ffc107; color: #333; padding: 3px 8px; border-radius: 3px;">⭐ Выделен</span>'
            )
        return ''
    team_highlight_badge.short_description = 'Выделен'

    # ⚡⚡⚡ СОХРАНЯЕМ REQUEST ДЛЯ ИСПОЛЬЗОВАНИЯ В ДРУГИХ МЕТОДАХ ⚡⚡⚡
    def get_queryset(self, request):
        self.request = request
        return super().get_queryset(request)

    # ⚡⚡⚡ МЕТОД ДЛЯ ОБРАБОТКИ СПИСКА (ТОЛЬКО ДЛЯ ТАБЛИЦЫ СТАТИСТИКИ) ⚡⚡⚡
    def changelist_view(self, request, extra_context=None):
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        # Если параметры есть в GET, сохраняем их в сессию
        if start_date and end_date:
            request.session['teacher_filter_start'] = start_date
            request.session['teacher_filter_end'] = end_date
        else:
            # Если параметров нет, пробуем взять из сессии
            start_date = request.session.get('teacher_filter_start')
            end_date = request.session.get('teacher_filter_end')

        print("\n" + "=" * 80)
        print("🔍 TEACHER ADMIN CHANGELIST VIEW")
        print(f"📅 start_date: {start_date}")
        print(f"📅 end_date: {end_date}")
        print("=" * 80)

        if start_date and end_date:
            try:
                from datetime import datetime
                start = datetime.strptime(start_date, '%Y-%m-%d').date()
                end = datetime.strptime(end_date, '%Y-%m-%d').date()

                print(f"\n✅ Период преобразован: {start} - {end}")

                extra_context = extra_context or {}
                teachers_data = []

                # Получаем всех учителей
                teachers = self.get_queryset(request)
                print(f"\n👥 Всего учителей: {teachers.count()}")

                for teacher in teachers:
                    print(f"\n{'─' * 50}")
                    print(f"👨‍🏫 Обработка учителя: {teacher.user.get_full_name()} (ID: {teacher.id})")

                    # Получаем статистику
                    earnings = teacher.get_teacher_earnings(start, end)

                    print(f"📊 earnings получены:")
                    print(f"   total_payments: {earnings.get('total_payments', 0)}")
                    print(f"   total_salaries: {earnings.get('total_salaries', 0)}")
                    print(f"   net_income: {earnings.get('net_income', 0)}")
                    print(f"   payments_count: {earnings.get('payments_count', 0)}")
                    print(f"   salaries_count: {earnings.get('salaries_count', 0)}")

                    teachers_data.append({
                        'teacher': teacher,
                        'earnings': earnings
                    })

                extra_context['teachers_data'] = teachers_data
                extra_context['start_date'] = start_date
                extra_context['end_date'] = end_date

                print(f"\n✅ teachers_data создан, размер: {len(teachers_data)}")

                # Выводим итоговые данные для всех учителей
                print(f"\n📊 ИТОГОВЫЕ ДАННЫЕ:")
                for i, item in enumerate(teachers_data):
                    teacher_name = item['teacher'].user.get_full_name()
                    earnings = item['earnings']
                    print(f"{i + 1}. {teacher_name}:")
                    print(f"   💰 total_payments: {earnings.get('total_payments', 0)}")
                    print(f"   💵 total_salaries: {earnings.get('total_salaries', 0)}")
                    print(f"   💹 net_income: {earnings.get('net_income', 0)}")

            except Exception as e:
                print(f"❌ ОШИБКА в changelist_view: {e}")
                import traceback
                traceback.print_exc()

        print("=" * 80 + "\n")
        return super().changelist_view(request, extra_context)

    actions = ['export_teachers_excel', 'calculate_payments',
               'show_on_team', 'hide_from_team', 'set_highlight', 'remove_highlight']

    def export_teachers_excel(self, request, queryset):
        """Экспорт выбранных учителей в Excel"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        from datetime import datetime
        from school.models import Lesson

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Учителя"

        headers = ['ID', 'Фамилия', 'Имя', 'Отчество', 'Email', 'Телефон',
                   'Предметы', 'Опыт', 'Всего уроков', 'Проведено уроков', 'Учеников']

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row, teacher in enumerate(queryset, start=2):
            subjects = ", ".join([s.name for s in teacher.subjects.all()])

            total_lessons = Lesson.objects.filter(teacher=teacher).count()
            completed_lessons = Lesson.objects.filter(teacher=teacher, status='completed').count()
            total_students = teacher.student_set.count()

            ws.cell(row=row, column=1, value=teacher.id)
            ws.cell(row=row, column=2, value=teacher.user.last_name)
            ws.cell(row=row, column=3, value=teacher.user.first_name)
            ws.cell(row=row, column=4, value=teacher.user.patronymic)
            ws.cell(row=row, column=5, value=teacher.user.email)
            ws.cell(row=row, column=6, value=teacher.user.phone)
            ws.cell(row=row, column=7, value=subjects)
            ws.cell(row=row, column=8, value=teacher.experience)
            ws.cell(row=row, column=9, value=total_lessons)
            ws.cell(row=row, column=10, value=completed_lessons)
            ws.cell(row=row, column=11, value=total_students)

        column_widths = [8, 15, 15, 15, 25, 15, 30, 8, 12, 15, 10]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"teachers_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response
    export_teachers_excel.short_description = "📥 Экспорт выбранных учителей в Excel"

    def calculate_payments(self, request, queryset):
        """Перейти к расчету выплат"""
        if queryset.count() == 1:
            teacher = queryset.first()
            return redirect(f'/admin/teacher-payments/?teacher_id={teacher.id}')
        else:
            self.message_user(request, 'Выберите одного учителя для расчета выплат', level='WARNING')
    calculate_payments.short_description = "💰 Расчет выплат"

    # НОВЫЕ ACTIONS
    def show_on_team(self, request, queryset):
        updated = queryset.update(show_on_team=True)
        self.message_user(request, f'✅ {updated} учителей будут показываться на странице команды')
    show_on_team.short_description = "✅ Показывать на странице команды"

    def hide_from_team(self, request, queryset):
        updated = queryset.update(show_on_team=False)
        self.message_user(request, f'❌ {updated} учителей скрыты со страницы команды')
    hide_from_team.short_description = "❌ Скрыть со страницы команды"

    def set_highlight(self, request, queryset):
        updated = queryset.update(team_highlight=True)
        self.message_user(request, f'⭐ {updated} учителей отмечены как "выделенные"')
    set_highlight.short_description = "⭐ Отметить как выделенных"

    def remove_highlight(self, request, queryset):
        updated = queryset.update(team_highlight=False)
        self.message_user(request, f'➖ У {updated} учителей снята отметка "выделенные"')
    remove_highlight.short_description = "➖ Снять выделение"


def export_teachers_excel(self, request, queryset):
    """Экспорт выбранных учителей в Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    from datetime import datetime
    from school.models import Lesson

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Учителя"

    # Заголовки (заменили Баланс кошелька на статистику)
    headers = ['ID', 'Фамилия', 'Имя', 'Отчество', 'Email', 'Телефон',
               'Предметы', 'Опыт', 'Всего уроков', 'Проведено уроков', 'Учеников']

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, teacher in enumerate(queryset, start=2):
        subjects = ", ".join([s.name for s in teacher.subjects.all()])

        # Считаем статистику учителя
        total_lessons = Lesson.objects.filter(teacher=teacher).count()
        completed_lessons = Lesson.objects.filter(teacher=teacher, status='completed').count()
        total_students = teacher.student_set.count()

        ws.cell(row=row, column=1, value=teacher.id)
        ws.cell(row=row, column=2, value=teacher.user.last_name)
        ws.cell(row=row, column=3, value=teacher.user.first_name)
        ws.cell(row=row, column=4, value=teacher.user.patronymic)
        ws.cell(row=row, column=5, value=teacher.user.email)
        ws.cell(row=row, column=6, value=teacher.user.phone)
        ws.cell(row=row, column=7, value=subjects)
        ws.cell(row=row, column=8, value=teacher.experience)
        ws.cell(row=row, column=9, value=total_lessons)  # Вместо баланса - всего уроков
        ws.cell(row=row, column=10, value=completed_lessons)  # Проведено уроков
        ws.cell(row=row, column=11, value=total_students)  # Учеников

    # Настройка ширины колонок (добавили одну колонку)
    column_widths = [8, 15, 15, 15, 25, 15, 30, 8, 12, 15, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"teachers_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


export_teachers_excel.short_description = "📥 Экспорт выбранных учителей в Excel"


def calculate_payments(self, request, queryset):
    """Перейти к расчету выплат"""
    if queryset.count() == 1:
        teacher = queryset.first()
        return redirect(f'/admin/school/teacher/{teacher.id}/payments/')
    else:
        self.message_user(request, 'Выберите одного учителя для расчета выплат', level='WARNING')
    calculate_payments.short_description = "💰 Расчет выплат"


# ==================== STUDENT ADMIN ====================

class StudentAdmin(admin.ModelAdmin):
    list_display = ('id', 'user_link', 'parent_name', 'parent_phone',
                    'get_teachers_count', 'last_lesson', 'balance_display', 'get_telegram_status')
    search_fields = ('user__first_name', 'user__last_name', 'user__email', 'parent_name')
    filter_horizontal = ('teachers',)
    list_filter = ('teachers',)
    raw_id_fields = ('user',)
    inlines = [StudentSubjectPriceInline]

    # Шаблон для кастомного отображения
    change_list_template = "admin/school/student/change_list.html"

    fieldsets = (
        (None, {
            'fields': ('user', 'teachers')
        }),
        ('Родители', {
            'fields': ('parent_name', 'parent_phone'),
        }),
        ('Заметки', {
            'fields': ('notes',),
            'classes': ('wide',),
        }),
    )
    def get_telegram_status(self, obj):
        if obj.user.telegram_chat_id:
            return f"✅ {obj.user.telegram_chat_id}"
        return "❌ Не подключен"
    get_telegram_status.short_description = "Telegram"

    def user_link(self, obj):
        url = f'/admin/school/user/{obj.user.id}/change/'
        return format_html('<a href="{}">{}</a>', url, obj.user.get_full_name())

    user_link.short_description = 'Ученик'

    def get_teachers_count(self, obj):
        return obj.teachers.count()

    get_teachers_count.short_description = 'Кол-во учителей'

    def last_lesson(self, obj):
        last = obj.lessons.order_by('-date').first()
        if last:
            return format_html('<a href="/admin/school/lesson/{}/change/">{} {}</a>',
                               last.id, last.date.strftime('%d.%m.%Y'), last.subject)
        return '-'

    last_lesson.short_description = 'Последний урок'

    def balance_display(self, obj):
        """Отображение баланса с цветом"""
        try:
            # Используем правильный расчет баланса
            balance = float(obj.user.balance_calculated)
        except (AttributeError, TypeError, ValueError):
            # Если что-то пошло не так
            from django.db.models import Sum
            from school.models import Payment, LessonAttendance

            total_deposits = Payment.objects.filter(
                user=obj.user,
                payment_type='income'
            ).aggregate(Sum('amount'))['amount__sum'] or 0

            attended_cost = LessonAttendance.objects.filter(
                student=obj,
                status='attended'
            ).aggregate(Sum('cost'))['cost__sum'] or 0

            balance = float(total_deposits - attended_cost)

        # Форматируем число
        balance_str = f"{balance:.2f} ₽"

        if balance < 0:
            return format_html('<span style="color: #dc3545; font-weight: bold;">{}</span>', balance_str)
        elif balance > 0:
            return format_html('<span style="color: #28a745; font-weight: bold;">{}</span>', balance_str)
        else:
            return format_html('<span style="color: #6c757d;">{}</span>', balance_str)

    balance_display.short_description = 'Баланс'


    # ⚡⚡⚡ МЕТОД ДЛЯ ОБРАБОТКИ СПИСКА СО СТАТИСТИКОЙ ⚡⚡⚡
    def changelist_view(self, request, extra_context=None):
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        # Сохраняем в сессию
        if start_date and end_date:
            request.session['student_filter_start'] = start_date
            request.session['student_filter_end'] = end_date
        else:
            start_date = request.session.get('student_filter_start')
            end_date = request.session.get('student_filter_end')

        print("\n" + "=" * 80)
        print("🔍 STUDENT ADMIN CHANGELIST VIEW")
        print(f"📅 start_date: {start_date}")
        print(f"📅 end_date: {end_date}")
        print("=" * 80)

        if start_date and end_date:
            try:
                from datetime import datetime
                start = datetime.strptime(start_date, '%Y-%m-%d').date()
                end = datetime.strptime(end_date, '%Y-%m-%d').date()

                print(f"\n✅ Период преобразован: {start} - {end}")

                extra_context = extra_context or {}
                students_data = []

                # Для подсчета итогов
                total_lessons = 0
                total_cost = 0
                total_balance = 0

                # Получаем всех учеников
                students = self.get_queryset(request)
                print(f"\n👥 Всего учеников: {students.count()}")

                for student in students:
                    print(f"\n{'─' * 50}")
                    print(f"👨‍🎓 Обработка ученика: {student.user.get_full_name()} (ID: {student.id})")

                    # Получаем статистику по урокам за период
                    from django.db.models import Sum, Count
                    from school.models import LessonAttendance

                    # Уроки за период со статусом 'attended'
                    attended_lessons = LessonAttendance.objects.filter(
                        student=student,
                        status='attended',
                        lesson__date__gte=start,
                        lesson__date__lte=end
                    )
                    total_deposits = Payment.objects.filter(
                        user=student.user,
                        payment_type='income'
                    ).aggregate(Sum('amount'))['amount__sum'] or 0

                    lessons_count = attended_lessons.count()
                    student_total_cost = attended_lessons.aggregate(Sum('cost'))['cost__sum'] or 0
                    student_balance = student.user.get_balance()
                    student_deposits_period = Payment.objects.filter(
                        user=student.user,
                        payment_type='income',
                        created_at__date__gte=start,
                        created_at__date__lte=end
                    ).aggregate(Sum('amount'))['amount__sum'] or 0

                    # Группировка по предметам
                    subjects_stats = attended_lessons.values(
                        'lesson__subject__name'
                    ).annotate(
                        count=Count('id'),
                        total=Sum('cost')
                    ).order_by('-total')

                    print(f"📊 Статистика:")
                    print(f"   уроков: {lessons_count}")
                    print(f"   сумма: {student_total_cost}")
                    for subj in subjects_stats:
                        print(f"   - {subj['lesson__subject__name']}: {subj['count']} ур. = {subj['total']}₽")

                    students_data.append({
                        'student': student,
                        'lessons_count': lessons_count,
                        'total_cost': student_total_cost,
                        'subjects_stats': subjects_stats,
                        'balance': student_balance,
                        'total_deposits': total_deposits,
                        'total_deposits_period': student_deposits_period,  # Пополнения за период
                    })

                    # Добавляем к итогам
                    total_lessons += lessons_count
                    total_cost += student_total_cost
                    total_balance += student_balance

                extra_context['students_data'] = students_data
                extra_context['start_date'] = start_date
                extra_context['end_date'] = end_date
                extra_context['total_lessons'] = total_lessons
                extra_context['total_cost'] = total_cost
                extra_context['total_balance'] = total_balance

                print(f"\n✅ students_data создан, размер: {len(students_data)}")
                print(f"📊 ИТОГО: уроков={total_lessons}, сумма={total_cost}, баланс={total_balance}")

            except Exception as e:
                print(f"❌ ОШИБКА в changelist_view: {e}")
                import traceback
                traceback.print_exc()

        print("=" * 80 + "\n")
        return super().changelist_view(request, extra_context)



    actions = ['export_students_excel', 'show_finance_report']

    def export_students_excel(self, request, queryset):
        """Экспорт выбранных учеников в Excel"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        from datetime import datetime

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ученики"

        headers = ['ID', 'Фамилия', 'Имя', 'Отчество', 'Email', 'Телефон',
                   'Родитель', 'Телефон родителя', 'Баланс', 'Учителя']

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row, student in enumerate(queryset, start=2):
            teachers = ", ".join([t.user.get_full_name() for t in student.teachers.all()[:3]])
            if student.teachers.count() > 3:
                teachers += f" и еще {student.teachers.count() - 3}"

            ws.cell(row=row, column=1, value=student.id)
            ws.cell(row=row, column=2, value=student.user.last_name)
            ws.cell(row=row, column=3, value=student.user.first_name)
            ws.cell(row=row, column=4, value=student.user.patronymic)
            ws.cell(row=row, column=5, value=student.user.email)
            ws.cell(row=row, column=6, value=student.user.phone)
            ws.cell(row=row, column=7, value=student.parent_name)
            ws.cell(row=row, column=8, value=student.parent_phone)
            ws.cell(row=row, column=9, value=float(student.user.get_balance()))
            ws.cell(row=row, column=10, value=teachers)

        column_widths = [8, 15, 15, 15, 25, 15, 20, 15, 12, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"students_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    export_students_excel.short_description = "📥 Экспорт выбранных учеников в Excel"

    def show_finance_report(self, request, queryset):
        """Показать финансовый отчет по ученикам"""
        if queryset.count() == 1:
            student = queryset.first()
            return redirect(f'/admin/school/student/{student.id}/report/')
        else:
            self.message_user(request, 'Выберите одного ученика для просмотра отчета', level='WARNING')

    show_finance_report.short_description = "📊 Финансовый отчет"


# ==================== LESSON FORMAT ADMIN ====================

class LessonFormatAdmin(admin.ModelAdmin):
    list_display = ('name', 'description', 'lessons_count')
    search_fields = ('name',)

    def lessons_count(self, obj):
        return obj.lesson_set.count()

    lessons_count.short_description = 'Занятий'


# ==================== LESSON ADMIN ====================

@admin.register(Lesson)
class LessonAdmin(admin.ModelAdmin):
    """
    Административный интерфейс для модели Lesson.

    Обеспечивает полное управление уроками, включая:
        * Проверку занятости учителя при создании
        * Отправку Telegram уведомлений о новых уроках
        * Финансовую статистику по каждому уроку
        * Календарное отображение
        * Массовые операции (завершение, отметка оплаты)

    Важные методы:
        save_model() - переопределяет сохранение с проверкой занятости
                       учителя и отправкой уведомлений
        _check_teacher_busy() - проверяет, не занят ли учитель в это время
        changelist_view() - переключает между таблицей и календарем
        mark_as_completed() - массовое завершение уроков
        export_lessons_finance() - экспорт финансовой статистики

    Кастомные действия:
        * Отметка уроков как проведенных
        * Отметка посещаемости как оплаченной
        * Отметка посещаемости как долга
        * Экспорт финансов
    """
    formfield_overrides = {
        models.TimeField: {'widget': forms.TimeInput(format='%H:%M', attrs={'type': 'time'})},
        models.DateField: {
            'widget': forms.DateInput(attrs={'type': 'date', 'value': datetime.now().strftime('%Y-%m-%d')})},
    }

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if obj is None:  # Только для создания нового урока
            from datetime import date, time
            form.base_fields['date'].initial = date.today()
            form.base_fields['start_time'].initial = time(9, 0)  # 09:00
            form.base_fields['end_time'].initial = time(10, 0)  # 10:00
        return form

    list_display = ('id', 'colored_subject', 'teacher_link', 'students_count', 'students_preview',
                    'date', 'start_time', 'status_badge', 'finance_preview')
    list_filter = ('status', 'subject', 'date', 'teacher', 'is_group')
    search_fields = ('teacher__user__last_name', 'students__user__last_name', 'subject__name')
    date_hierarchy = 'date'
    raw_id_fields = ('teacher',)
    inlines = [LessonAttendanceInline]
    readonly_fields = ('finance_stats', 'video_room')
    change_form_template = "admin/school/lesson/change_form.html"

    fieldsets = (
        ('Основная информация', {
            'fields': ('teacher', 'subject', 'format', 'is_group')
        }),
        ('Время', {
            'fields': ('date', 'start_time', 'end_time', 'duration')
        }),
        ('Финансы', {
            'fields': ('price_type', 'base_cost', 'base_teacher_payment'),
            'classes': ('wide',),
            'description': 'Базовая стоимость урока. Индивидуальные цены можно настроить в таблице посещаемости ниже.'
        }),
        ('Финансовая статистика', {
            'fields': ('finance_stats',),
            'classes': ('wide', 'collapse'),
        }),
        ('Платформа', {
            'fields': ('meeting_link', 'meeting_platform', 'video_room')
        }),
        ('Статус и заметки', {
            'fields': ('status', 'notes')
        }),
    )

    def teacher_link(self, obj):
        url = f'/admin/school/teacher/{obj.teacher.id}/change/'
        return format_html('<a href="{}">{}</a>', url, obj.teacher.user.get_full_name())

    teacher_link.short_description = 'Учитель'

    def colored_subject(self, obj):
        colors = {
            'scheduled': '#007bff',
            'completed': '#28a745',
            'cancelled': '#dc3545',
            'overdue': '#fd7e14',
        }
        color = colors.get(obj.status, '#6c757d')
        return format_html('<span style="color: {}; font-weight: bold;">{}</span>', color, obj.subject.name)

    colored_subject.short_description = 'Предмет'

    def students_count(self, obj):
        count = obj.students.count()
        if count == 0:
            return format_html('<span style="color: #dc3545;">❌ 0</span>')
        elif count == 1:
            return format_html('👤 1')
        else:
            return format_html('👥 {}', count)

    students_count.short_description = 'Кол-во'

    def students_preview(self, obj):
        students = obj.students.all()
        if not students:
            return format_html('<span style="color: #dc3545;">❌ Нет учеников</span>')

        names = []
        for student in students:
            last_name = student.user.last_name or ''
            first_name = student.user.first_name or ''

            if last_name and first_name:
                names.append(f"{last_name} {first_name}")
            elif last_name:
                names.append(last_name)
            elif first_name:
                names.append(first_name)
            else:
                names.append(student.user.username)

        return ", ".join(names)

    students_preview.short_description = 'Ученики'

    def status_badge(self, obj):
        status_colors = {
            'scheduled': ('#007bff', 'Запланировано'),
            'completed': ('#28a745', 'Проведено'),
            'cancelled': ('#dc3545', 'Отменено'),
            'overdue': ('#fd7e14', 'Просрочено'),
        }
        color, text = status_colors.get(obj.status, ('#6c757d', obj.status))
        return format_html(
            '<span style="background: {}; color: white; padding: 3px 8px; border-radius: 3px;">{}</span>',
            color, text)

    status_badge.short_description = 'Статус'

    def finance_preview(self, obj):
        calculator = LessonFinanceCalculator(obj)
        stats = calculator.stats

        if stats['students_total'] == 0:
            return '-'

        return format_html(
            '<span title="Всего: {}₽\nВыплата: {}₽\nПрисутствовало: {}/{}">💰 {}₽</span>',
            stats['total_cost'], stats['teacher_payment'],
            stats['students_attended'], stats['students_total'],
            stats['total_cost']
        )

    finance_preview.short_description = 'Финансы'

    def finance_stats(self, obj):
        calculator = LessonFinanceCalculator(obj)
        stats = calculator.stats

        html = f"""
        <div style="background: #f8f9fa; padding: 15px; border-radius: 5px;">
            <h3 style="margin-top: 0;">Финансовая статистика урока</h3>
            <table style="width: 100%; border-collapse: collapse;">
                <tr>
                    <td style="padding: 5px;"><strong>Всего учеников:</strong></td>
                    <td style="padding: 5px;">{stats['students_total']}</td>
                </tr>
                <tr>
                    <td style="padding: 5px;"><strong>Присутствовало:</strong></td>
                    <td style="padding: 5px; color: #28a745;">{stats['students_attended']}</td>
                </tr>
                <tr>
                    <td style="padding: 5px;"><strong>В долг:</strong></td>
                    <td style="padding: 5px; color: #dc3545;">{stats['students_debt']}</td>
                </tr>
                <tr>
                    <td style="padding: 5px;"><strong>Отсутствовало:</strong></td>
                    <td style="padding: 5px; color: #6c757d;">{stats['students_absent']}</td>
                </tr>
                <tr>
                    <td style="padding: 5px;"><strong>Общая стоимость:</strong></td>
                    <td style="padding: 5px; font-weight: bold;">{stats['total_cost']} ₽</td>
                </tr>
                <tr>
                    <td style="padding: 5px;"><strong>Выплата учителю:</strong></td>
                    <td style="padding: 5px; font-weight: bold;">{stats['teacher_payment']} ₽</td>
                </tr>
                <tr>
                    <td style="padding: 5px;"><strong>Стоимость присутствовавших:</strong></td>
                    <td style="padding: 5px;">{stats['attended_cost']} ₽</td>
                </tr>
                <tr>
                    <td style="padding: 5px;"><strong>Задолженность:</strong></td>
                    <td style="padding: 5px; color: #dc3545;">{stats['debt_cost']} ₽</td>
                </tr>
            </table>
        </div>
        """
        return format_html(html)

    finance_stats.short_description = 'Финансовая статистика'

    def has_report(self, obj):
        if hasattr(obj, 'report'):
            url = f'/admin/school/lessonreport/{obj.report.id}/change/'
            return format_html('<a href="{}" style="color: #28a745;">✅ Отчет #{}</a>', url, obj.report.id)
        return '❌ Нет отчета'

    has_report.short_description = 'Отчет'

    actions = ['export_lessons_finance', 'mark_as_completed', 'mark_as_paid', 'mark_as_debt']

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('<int:lesson_id>/complete/',
                 self.admin_site.admin_view(admin_complete_lesson),
                 name='complete-lesson'),
            path('bulk-complete/',
                 self.admin_site.admin_view(self.bulk_complete_view),
                 name='bulk-complete-lessons'),
        ]
        return custom_urls + urls

    def bulk_complete_view(self, request):
        lesson_ids = request.GET.getlist('ids')
        lessons = Lesson.objects.filter(id__in=lesson_ids)

        if request.method == 'POST':
            from django.contrib import messages
            completed = 0
            for lesson in lessons:
                if lesson.status == 'scheduled':
                    completed += 1
            messages.success(request, f'✅ Завершено {completed} уроков')
            return redirect('admin:school_lesson_changelist')

        context = {
            'lessons': lessons,
            'title': 'Массовое завершение уроков',
        }
        return render(request, 'admin/school/lesson/bulk_complete.html', context)

    # ✅ ЕДИНСТВЕННЫЙ МЕТОД СОХРАНЕНИЯ (с проверкой и Telegram)
    def save_model(self, request, obj, form, change):
        """Сохраняет урок с проверкой на занятость учителя и отправляет уведомление"""

        # Проверка на занятость учителя
        try:
            self._check_teacher_busy(obj)
        except ValidationError as e:
            form.add_error(None, e)
            return

        is_new = obj.pk is None
        super().save_model(request, obj, form, change)

        # Отправляем уведомление только для новых уроков
        if is_new:
            try:
                from school.telegram import notify_new_lesson
                notify_new_lesson(obj)
            except Exception as e:
                print(f"❌ Ошибка отправки Telegram уведомления: {e}")

    def _check_teacher_busy(self, lesson):
        """Проверяет, свободен ли учитель в это время"""
        from datetime import datetime

        existing_lessons = Lesson.objects.filter(
            teacher=lesson.teacher,
            date=lesson.date,
            status__in=['scheduled', 'completed']
        ).exclude(pk=lesson.pk)

        lesson_start = datetime.combine(lesson.date, lesson.start_time)
        lesson_end = datetime.combine(lesson.date, lesson.end_time)

        for existing in existing_lessons:
            existing_start = datetime.combine(existing.date, existing.start_time)
            existing_end = datetime.combine(existing.date, existing.end_time)

            if lesson_start < existing_end and lesson_end > existing_start:
                raise ValidationError(
                    f'Учитель {lesson.teacher.user.get_full_name()} уже занят в это время! '
                    f'Существующий урок: {existing.start_time} - {existing.end_time}'
                )

    # ✅ МЕТОД ДЛЯ КАЛЕНДАРЯ
    def changelist_view(self, request, extra_context=None):
        if request.GET.get('view') == 'calendar':
            lessons = self.get_queryset(request).select_related(
                'teacher__user', 'subject'
            ).prefetch_related(
                Prefetch(
                    'attendance',
                    queryset=LessonAttendance.objects.select_related('student__user')
                )
            )

            calendar_events = []
            for lesson in lessons:
                calculator = LessonFinanceCalculator(lesson)
                stats = calculator.stats

                subject_short = lesson.subject.name[:4]
                teacher_last = lesson.teacher.user.last_name

                if stats['students_total'] == 0:
                    students_text = "нет"
                elif stats['students_total'] == 1:
                    student = lesson.attendance.first().student
                    students_text = student.user.first_name
                else:
                    students_text = f"{stats['students_total']} уч."

                title = f"{subject_short} {teacher_last} - {students_text}"

                status_colors = {
                    'completed': '#28a745',
                    'cancelled': '#dc3545',
                    'overdue': '#fd7e14',
                    'scheduled': '#007bff',
                }
                bg_color = status_colors.get(lesson.status, '#6c757d')

                calendar_events.append({
                    'title': title,
                    'start': f"{lesson.date}T{lesson.start_time}",
                    'end': f"{lesson.date}T{lesson.end_time}",
                    'url': f"/admin/school/lesson/{lesson.id}/change/",
                    'backgroundColor': bg_color,
                    'borderColor': bg_color,
                    'textColor': 'white',
                    'finance': {
                        'total_cost': stats['total_cost'],
                        'teacher_payment': stats['teacher_payment']
                    }
                })

            extra_context = extra_context or {}
            extra_context['calendar_events'] = calendar_events
            extra_context['title'] = 'Календарь занятий'

            return render(request, 'admin/school/lesson/change_list_calendar.html', extra_context)

        return super().changelist_view(request, extra_context)

    def response_change(self, request, obj):
        if "_complete-lesson" in request.POST:
            return redirect('admin:complete-lesson', lesson_id=obj.id)
        return super().response_change(request, obj)

    def export_lessons_finance(self, request, queryset):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Финансы уроков"

        headers = ['ID', 'Дата', 'Учитель', 'Предмет', 'Учеников',
                   'Присутствовало', 'В долг', 'Общая стоимость', 'Выплата учителю']

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row, lesson in enumerate(queryset, start=2):
            calculator = LessonFinanceCalculator(lesson)
            stats = calculator.stats

            ws.cell(row=row, column=1, value=lesson.id)
            ws.cell(row=row, column=2, value=lesson.date.strftime('%d.%m.%Y'))
            ws.cell(row=row, column=3, value=lesson.teacher.user.get_full_name())
            ws.cell(row=row, column=4, value=lesson.subject.name)
            ws.cell(row=row, column=5, value=stats['students_total'])
            ws.cell(row=row, column=6, value=stats['students_attended'])
            ws.cell(row=row, column=7, value=stats['students_debt'])
            ws.cell(row=row, column=8, value=stats['total_cost'])
            ws.cell(row=row, column=9, value=stats['teacher_payment'])

        column_widths = [8, 12, 25, 20, 8, 10, 8, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"lessons_finance_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    export_lessons_finance.short_description = "📊 Экспорт финансов уроков"

    def mark_as_completed(self, request, queryset):
        from django.contrib import messages

        completed = 0
        for lesson in queryset.filter(status='scheduled'):
            if lesson.attendance.exists():
                lesson.status = 'completed'
                lesson.save()
                lesson.attendance.update(status='attended')
                completed += 1

        self.message_user(request, f'✅ {completed} уроков отмечены как проведенные')

    mark_as_completed.short_description = "✅ Отметить как проведенные"

    def mark_as_paid(self, request, queryset):
        updated = LessonAttendance.objects.filter(
            lesson__in=queryset,
            status='debt'
        ).update(status='attended')
        self.message_user(request, f'💰 {updated} записей отмечены как оплаченные')

    mark_as_paid.short_description = "💰 Отметить как оплаченные"

    def mark_as_debt(self, request, queryset):
        updated = LessonAttendance.objects.filter(
            lesson__in=queryset,
            status='attended'
        ).update(status='debt')
        self.message_user(request, f'⚠️ {updated} записей отмечены как долг')

    mark_as_debt.short_description = "⚠️ Отметить как долг"

# ==================== LESSON REPORT ADMIN ====================

@admin.register(LessonReport)
class LessonReportAdmin(admin.ModelAdmin):
    list_display = ('id', 'lesson_link', 'topic_preview', 'created_at')
    list_filter = ('created_at', 'lesson__subject')
    search_fields = ('topic', 'lesson__subject__name')
    readonly_fields = ('created_at',)
    raw_id_fields = ('lesson',)

    def lesson_link(self, obj):
        url = f'/admin/school/lesson/{obj.lesson.id}/change/'
        return format_html('<a href="{}">{} #{}</a>', url, obj.lesson.subject, obj.lesson.id)

    lesson_link.short_description = 'Занятие'

    def topic_preview(self, obj):
        return obj.topic[:50] + '...' if len(obj.topic) > 50 else obj.topic

    topic_preview.short_description = 'Тема'


# ==================== PAYMENT ADMIN ====================

@admin.register(Payment)
class PaymentAdmin(admin.ModelAdmin):
    """
    Административный интерфейс для модели Payment.

    Критически важный класс для финансового учета.
    Автоматически корректирует балансы пользователей при
    создании, изменении и удалении платежей.

    Ключевые методы:
        save_model() - при создании/изменении платежа обновляет:
            * Баланс ученика (income/expense)
            * Кошелек учителя (teacher_payment/teacher_salary)
            * Создает уведомления для пользователей

        delete_model() - при удалении платежа корректирует баланс
                         в обратную сторону

        delete_queryset() - массовое удаление с корректировкой балансов

    Кастомные действия:
        export_payments_excel() - экспорт платежей в Excel

    Важно: Все финансовые операции проходят через этот класс,
    обеспечивая консистентность данных.
    """
    list_display = ('id', 'user_link', 'amount_colored', 'payment_type_badge',
                    'description', 'lesson_link', 'created_at')
    list_filter = ('payment_type', 'created_at')
    search_fields = ('user__first_name', 'user__last_name', 'description')
    date_hierarchy = 'created_at'
    raw_id_fields = ('user', 'lesson')
    readonly_fields = ('created_at',)

    fieldsets = (
        (None, {
            'fields': ('user', 'payment_type', 'amount', 'description')
        }),
        ('Связи', {
            'fields': ('lesson',),
            'classes': ('wide',),
        }),
        ('Дата', {
            'fields': ('created_at',),
        }),
    )

    def user_link(self, obj):
        url = f'/admin/school/user/{obj.user.id}/change/'
        return format_html('<a href="{}">{}</a>', url, obj.user.get_full_name())

    user_link.short_description = 'Пользователь'

    def amount_colored(self, obj):
        if obj.payment_type == 'income':
            return format_html('<span style="color: #28a745;">+{} ₽</span>', obj.amount)
        elif obj.payment_type == 'expense':
            return format_html('<span style="color: #dc3545;">-{} ₽</span>', obj.amount)
        else:
            return format_html('<span style="color: #17a2b8;">{} ₽</span>', obj.amount)

    amount_colored.short_description = 'Сумма'

    def payment_type_badge(self, obj):
        type_colors = {
            'income': ('#28a745', 'Пополнение'),
            'expense': ('#dc3545', 'Списание'),
            # 'teacher_payment': ('#17a2b8', 'Начисление учителю'),
            'teacher_salary': ('#ffc107', 'Зарплата учителя'),
        }
        color, text = type_colors.get(obj.payment_type, ('#6c757d', obj.payment_type))
        return format_html(
            '<span style="background: {}; color: white; padding: 3px 8px; border-radius: 3px;">{}</span>',
            color, text)

    payment_type_badge.short_description = 'Тип'

    def lesson_link(self, obj):
        if obj.lesson:
            url = f'/admin/school/lesson/{obj.lesson.id}/change/'
            return format_html('<a href="{}">Урок #{}</a>', url, obj.lesson.id)
        return '-'

    lesson_link.short_description = 'Урок'

    # ⚡⚡⚡ МЕТОД ДЛЯ СОХРАНЕНИЯ (ОБНОВЛЕНИЕ БАЛАНСА) ⚡⚡⚡
    def save_model(self, request, obj, form, change):
        """Сохраняет платеж и обновляет баланс пользователя"""
        # Сначала сохраняем платеж
        super().save_model(request, obj, form, change)

        # Обновляем баланс в зависимости от типа платежа
        if obj.payment_type == 'income':
            # Пополнение счета ученика
            obj.user.balance += obj.amount
            obj.user.save()

            # ✅ Уведомление ТОЛЬКО ученику о пополнении
            Notification.objects.create(
                user=obj.user,
                title='💰 Пополнение баланса',
                message=f'Ваш баланс пополнен на {obj.amount} ₽',
                notification_type='payment_received',
                link='/student/dashboard/',
                payment=obj,
            )

        elif obj.payment_type == 'expense':
            # Списание со счета ученика
            obj.user.balance -= obj.amount
            obj.user.save()

            # Уведомление о списании
            Notification.objects.create(
                user=obj.user,
                title='💸 Списание средств',
                message=f'С вашего баланса списано {obj.amount} ₽',
                notification_type='payment_withdrawn',
                link='/student/dashboard/'
            )

        elif obj.payment_type == 'teacher_payment':
            # Выплата учителю
            try:
                teacher = obj.user.teacher_profile
                teacher.wallet_balance += obj.amount
                teacher.save()

                # ✅ Уведомление учителю о выплате
                Notification.objects.create(
                    user=obj.user,
                    title='💰 Выплата начислена',
                    message=f'Вам начислена выплата {obj.amount} ₽',
                    notification_type='payment_received',
                    link='/teacher/dashboard/'
                )
            except Teacher.DoesNotExist:
                print(f"⚠️ Пользователь {obj.user.username} не является учителем")

        elif obj.payment_type == 'teacher_salary':
            # Зарплата учителя
            try:
                teacher = obj.user.teacher_profile
                teacher.wallet_balance += obj.amount
                teacher.save()

                # ✅ Уведомление учителю о зарплате
                Notification.objects.create(
                    user=obj.user,
                    title='💰 Зарплата начислена',
                    message=f'Вам начислена зарплата {obj.amount} ₽',
                    notification_type='payment_received',
                    link='/teacher/dashboard/'
                )
            except Teacher.DoesNotExist:
                print(f"⚠️ Пользователь {obj.user.username} не является учителем")

    # ⚡⚡⚡ МЕТОД ДЛЯ УДАЛЕНИЯ ОДНОГО ПЛАТЕЖА ⚡⚡⚡
    def delete_model(self, request, obj):
        """При удалении платежа корректируем баланс"""
        # Запоминаем данные до удаления
        user = obj.user
        amount = obj.amount
        payment_type = obj.payment_type
        description = obj.description

        # Корректируем баланс в зависимости от типа платежа
        if payment_type == 'income':
            # Если удаляем пополнение - уменьшаем баланс
            user.balance -= amount
            user.save()

            # Уведомление об удалении пополнения
            Notification.objects.create(
                user=user,
                title='⚠️ Пополнение удалено',
                message=f'Пополнение на {amount} ₽ "{description}" было удалено. Баланс скорректирован.',
                notification_type='system',
            )

        elif payment_type == 'expense':
            # Если удаляем списание - увеличиваем баланс (возвращаем деньги)
            user.balance += amount
            user.save()

            # Уведомление об удалении списания
            Notification.objects.create(
                user=user,
                title='⚠️ Списание отменено',
                message=f'Списание {amount} ₽ "{description}" отменено. Деньги возвращены на баланс.',
                notification_type='system',
            )

        elif payment_type == 'teacher_payment':
            # Для выплаты учителю - уменьшаем wallet_balance
            try:
                teacher = user.teacher_profile
                teacher.wallet_balance -= amount
                teacher.save()

                # Уведомление учителю
                Notification.objects.create(
                    user=user,
                    title='⚠️ Выплата отменена',
                    message=f'Выплата {amount} ₽ "{description}" была отменена. Баланс кошелька скорректирован.',
                    notification_type='system',
                )
            except Teacher.DoesNotExist:
                pass

        # Удаляем сам платеж
        super().delete_model(request, obj)

        # Добавляем сообщение в админку
        messages.success(request, f'✅ Платеж удален. Баланс пользователя {user.username} скорректирован.')


# ⚡⚡⚡ ИСПРАВЛЕННЫЙ МЕТОД ДЛЯ МАССОВОГО УДАЛЕНИЯ ⚡⚡⚡
def delete_queryset(self, request, queryset):
    """При массовом удалении корректируем балансы"""
    count = queryset.count()

    with transaction.atomic():
        for obj in queryset:
            # Для каждого платежа применяем ту же логику
            user = obj.user
            amount = obj.amount
            payment_type = obj.payment_type

            if payment_type == 'income':
                user.balance -= amount
                user.save()
            elif payment_type == 'expense':
                user.balance += amount
                user.save()
            elif payment_type == 'teacher_payment':
                try:
                    teacher = user.teacher_profile
                    teacher.wallet_balance -= amount
                    teacher.save()
                except Teacher.DoesNotExist:
                    pass

        # Удаляем все платежи разом
        super().delete_queryset(request, queryset)

    # ✅ ИСПРАВЛЕНО: передаем request в messages
    self.message_user(request, f'✅ Удалено {count} платежей. Балансы пользователей скорректированы.', level='SUCCESS')
    actions = ['export_payments_excel']

    def export_payments_excel(self, request, queryset):
        """Экспорт платежей в Excel"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        from datetime import datetime

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Платежи"

        headers = ['ID', 'Дата', 'Пользователь', 'Тип', 'Сумма', 'Описание']

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row, payment in enumerate(queryset, start=2):
            ws.cell(row=row, column=1, value=payment.id)
            ws.cell(row=row, column=2, value=payment.created_at.strftime('%d.%m.%Y %H:%M'))
            ws.cell(row=row, column=3, value=payment.user.get_full_name())
            ws.cell(row=row, column=4, value=payment.get_payment_type_display())
            ws.cell(row=row, column=5, value=float(payment.amount))
            ws.cell(row=row, column=6, value=payment.description)

        column_widths = [8, 20, 30, 15, 12, 40]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"payments_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    export_payments_excel.short_description = "📥 Экспорт платежей в Excel"


# ==================== SCHEDULE ADMIN ====================

# @admin.register(Schedule)
class ScheduleAdmin(admin.ModelAdmin):
    formfield_overrides = {
        models.TimeField: {'widget': forms.TimeInput(format='%H:%M', attrs={'type': 'time'})},
        models.DateField: {'widget': forms.DateInput(attrs={'type': 'date'})},
    }

    list_display = ('teacher', 'date', 'day_of_week_display', 'start_time', 'end_time', 'is_active')
    list_filter = ('date', 'is_active', 'teacher')
    search_fields = ('teacher__user__last_name',)

    def day_of_week_display(self, obj):
        days = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
        return days[obj.date.weekday()]

    day_of_week_display.short_description = 'День'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('calendar-data/', self.admin_site.admin_view(schedule_calendar_data), name='schedule-calendar-data'),
        ]
        return custom_urls + urls


# ==================== TRIAL REQUEST ADMIN ====================

@admin.register(TrialRequest)
class TrialRequestAdmin(admin.ModelAdmin):
    list_display = ('name', 'phone', 'email', 'subject', 'created_at', 'is_processed_badge')
    list_filter = ('is_processed', 'subject', 'created_at')
    search_fields = ('name', 'phone', 'email')
    date_hierarchy = 'created_at'
    actions = ['mark_as_processed']

    def is_processed_badge(self, obj):
        if obj.is_processed:
            return format_html('<span style="color: #28a745;">✅ Обработано</span>')
        return format_html('<span style="color: #ffc107;">⏳ Новое</span>')

    is_processed_badge.short_description = 'Статус'

    def mark_as_processed(self, request, queryset):
        updated = queryset.update(is_processed=True)
        self.message_user(request, f'✅ {updated} заявок отмечены как обработанные')

    mark_as_processed.short_description = "✅ Отметить как обработанные"


# ==================== NOTIFICATION ADMIN ====================

@admin.register(Notification)
class NotificationAdmin(admin.ModelAdmin):
    list_display = ('id', 'user', 'title', 'notification_type_icon', 'is_read_badge', 'created_at')
    list_filter = ('notification_type', 'is_read', 'created_at')
    search_fields = ('user__username', 'user__email', 'title')
    date_hierarchy = 'created_at'
    raw_id_fields = ('user',)
    list_per_page = 50

    def notification_type_icon(self, obj):
        icons = {
            'lesson_reminder': '🔔',
            'lesson_canceled': '❌',
            'lesson_completed': '✅',
            'payment_received': '💰',
            'payment_withdrawn': '💸',
            'material_added': '📚',
            'homework_assigned': '📝',
            'feedback_received': '⭐',
            'system': '⚙',
        }
        icon = icons.get(obj.notification_type, '📢')
        return f"{icon} {obj.get_notification_type_display()}"

    notification_type_icon.short_description = 'Тип'

    def is_read_badge(self, obj):
        if obj.is_read:
            return format_html('<span style="color: #6c757d;">✓ Прочитано</span>')
        return format_html('<span style="color: #007bff; font-weight: bold;">● Новое</span>')

    is_read_badge.short_description = 'Статус'

    actions = ['mark_as_read', 'mark_as_unread']

    def mark_as_read(self, request, queryset):
        updated = queryset.update(is_read=True)
        self.message_user(request, f'✅ {updated} уведомлений отмечены как прочитанные')

    mark_as_read.short_description = "✓ Отметить как прочитанные"

    def mark_as_unread(self, request, queryset):
        updated = queryset.update(is_read=False)
        self.message_user(request, f'🔄 {updated} уведомлений отмечены как непрочитанные')

    mark_as_unread.short_description = "🔄 Отметить как непрочитанные"


# ==================== LESSON FEEDBACK ADMIN ====================

@admin.register(LessonFeedback)
class LessonFeedbackAdmin(admin.ModelAdmin):
    list_display = ('id', 'lesson_link', 'student', 'teacher', 'rating_stars',
                    'comment_preview', 'is_public_badge', 'created_at')
    list_filter = ('rating', 'is_public', 'created_at')
    search_fields = ('student__user__last_name', 'teacher__user__last_name', 'comment')
    raw_id_fields = ('lesson', 'student', 'teacher')
    date_hierarchy = 'created_at'
    actions = ['make_public', 'make_private']

    def lesson_link(self, obj):
        url = f'/admin/school/lesson/{obj.lesson.id}/change/'
        return format_html('<a href="{}">{} #{}</a>', url, obj.lesson.subject, obj.lesson.id)

    lesson_link.short_description = 'Урок'

    def rating_stars(self, obj):
        return '⭐' * obj.rating

    rating_stars.short_description = 'Оценка'

    def comment_preview(self, obj):
        if obj.comment:
            return obj.comment[:50] + '...' if len(obj.comment) > 50 else obj.comment
        return '-'

    comment_preview.short_description = 'Комментарий'

    def is_public_badge(self, obj):
        if obj.is_public:
            return format_html('<span style="color: #28a745;">✅ Публичный</span>')
        return format_html('<span style="color: #6c757d;">🔒 Приватный</span>')

    is_public_badge.short_description = 'Видимость'

    def make_public(self, request, queryset):
        updated = queryset.update(is_public=True)
        self.message_user(request, f'✅ {updated} оценок опубликовано')

    make_public.short_description = '✅ Опубликовать'

    def make_private(self, request, queryset):
        updated = queryset.update(is_public=False)
        self.message_user(request, f'🔒 {updated} оценок скрыто')

    make_private.short_description = '🔒 Скрыть'


# ==================== TEACHER RATING ADMIN ====================

@admin.register(TeacherRating)
class TeacherRatingAdmin(admin.ModelAdmin):
    list_display = ('teacher', 'average_rating_display', 'total_feedbacks',
                    'rating_distribution', 'updated_at')
    list_select_related = ('teacher__user',)
    readonly_fields = ('teacher', 'average_rating', 'total_feedbacks',
                       'rating_5_count', 'rating_4_count', 'rating_3_count',
                       'rating_2_count', 'rating_1_count', 'updated_at')

    def average_rating_display(self, obj):
        stars = '⭐' * int(obj.average_rating)
        return f"{stars} ({obj.average_rating:.1f})"

    average_rating_display.short_description = 'Средний балл'

    def rating_distribution(self, obj):
        total = obj.total_feedbacks or 1
        html = '<div style="width: 200px;">'
        for rating, count in [(5, obj.rating_5_count), (4, obj.rating_4_count),
                              (3, obj.rating_3_count), (2, obj.rating_2_count),
                              (1, obj.rating_1_count)]:
            width = (count / total) * 100
            html += f'<div style="margin: 2px 0;">{rating}⭐: <span style="display: inline-block; width: {width}px; height: 10px; background: #ffc107;"></span> {count}</div>'
        html += '</div>'
        return format_html(html)

    rating_distribution.short_description = 'Распределение'


# ==================== HOMEWORK ADMIN ====================

@admin.register(Homework)
class HomeworkAdmin(admin.ModelAdmin):
    list_display = ('id', 'colored_title', 'student_link', 'teacher_link',
                    'subject', 'deadline_colored', 'status_badge')
    list_filter = ('subject', 'is_active', 'deadline')
    search_fields = ('title', 'student__user__last_name', 'teacher__user__last_name')
    date_hierarchy = 'deadline'
    raw_id_fields = ('student', 'teacher', 'subject')
    list_per_page = 25
    save_on_top = True

    def student_link(self, obj):
        url = f'/admin/school/student/{obj.student.id}/change/'
        return format_html('<a href="{}">{}</a>', url, obj.student.user.get_full_name())

    student_link.short_description = 'Ученик'

    def teacher_link(self, obj):
        url = f'/admin/school/teacher/{obj.teacher.id}/change/'
        return format_html('<a href="{}">{}</a>', url, obj.teacher.user.get_full_name())

    teacher_link.short_description = 'Учитель'

    def colored_title(self, obj):
        return format_html('<span style="color: #2c3e50; font-weight: bold;">{}</span>', obj.title)

    colored_title.short_description = 'Название'

    def deadline_colored(self, obj):
        now = timezone.now()
        if obj.deadline < now:
            return format_html('<span style="color: #dc3545;">⚠️ {}</span>',
                               obj.deadline.strftime('%d.%m.%Y %H:%M'))
        elif (obj.deadline - now).days < 1:
            return format_html('<span style="color: #ffc107;">⚡ {}</span>',
                               obj.deadline.strftime('%d.%m.%Y %H:%M'))
        else:
            return format_html('<span style="color: #28a745;">✅ {}</span>',
                               obj.deadline.strftime('%d.%m.%Y %H:%M'))

    deadline_colored.short_description = 'Срок сдачи'

    def status_badge(self, obj):
        status = obj.get_status()
        status_colors = {
            'pending': ('#ffc107', '⏳ Ожидает'),
            'submitted': ('#17a2b8', '📤 На проверке'),
            'checked': ('#28a745', '✅ Проверено'),
            'overdue': ('#dc3545', '⚠️ Просрочено'),
        }
        color, text = status_colors.get(status, ('#6c757d', '❓'))
        return format_html(
            '<span style="background: {}; color: white; padding: 3px 8px; border-radius: 3px;">{}</span>',
            color, text)

    status_badge.short_description = 'Статус'


# ==================== HOMEWORK SUBMISSION ADMIN ====================

@admin.register(HomeworkSubmission)
class HomeworkSubmissionAdmin(admin.ModelAdmin):
    list_display = ('id', 'homework_link', 'student_name', 'submitted_at',
                    'status_colored', 'grade_display')
    list_filter = ('status', 'submitted_at')
    search_fields = ('homework__title', 'student__user__last_name')
    date_hierarchy = 'submitted_at'
    raw_id_fields = ('homework', 'student')

    def homework_link(self, obj):
        url = f'/admin/school/homework/{obj.homework.id}/change/'
        return format_html('<a href="{}">{}</a>', url, obj.homework.title)

    homework_link.short_description = 'Задание'

    def student_name(self, obj):
        return obj.student.user.get_full_name()

    student_name.short_description = 'Ученик'

    def status_colored(self, obj):
        if obj.status == 'submitted':
            return format_html('<span style="color: #17a2b8;">📤 Ожидает проверки</span>')
        return format_html('<span style="color: #28a745;">✅ Проверено</span>')

    status_colored.short_description = 'Статус'

    def grade_display(self, obj):
        if obj.grade:
            return format_html(
                '<span style="background: #28a745; color: white; padding: 2px 6px; border-radius: 3px;">{}/5</span>',
                obj.grade)
        return '—'

    grade_display.short_description = 'Оценка'


# ==================== GROUP LESSON ADMIN ====================

@admin.register(GroupLesson)
class GroupLessonAdmin(admin.ModelAdmin):
    list_display = ('id', 'subject', 'teacher', 'date', 'start_time',
                    'students_count', 'status_badge', 'finance_preview')
    list_filter = ('status', 'subject', 'teacher', 'date')
    search_fields = ('subject__name', 'teacher__user__last_name', 'notes')
    inlines = [GroupEnrollmentInline]

    fieldsets = (
        ('Основное', {
            'fields': ('teacher', 'subject', 'format', 'date', 'start_time', 'end_time')
        }),
        ('Финансы', {
            'fields': ('price_type', 'base_price', 'teacher_payment'),
        }),
        ('Платформа', {
            'fields': ('meeting_link', 'meeting_platform', 'video_room')
        }),
        ('Статус и заметки', {
            'fields': ('status', 'notes')
        }),
    )

    def students_count(self, obj):
        return obj.enrollments.count()

    students_count.short_description = 'Учеников'

    def status_badge(self, obj):
        status_colors = {
            'scheduled': ('#007bff', 'Запланировано'),
            'completed': ('#28a745', 'Проведено'),
            'cancelled': ('#dc3545', 'Отменено'),
        }
        color, text = status_colors.get(obj.status, ('#6c757d', obj.status))
        return format_html(
            '<span style="background: {}; color: white; padding: 3px 8px; border-radius: 3px;">{}</span>',
            color, text)

    status_badge.short_description = 'Статус'

    def finance_preview(self, obj):
        return format_html(
            '<span title="Сбор: {}₽\nВыплата: {}₽">💰 {}₽</span>',
            obj.get_total_cost(), obj.teacher_payment, obj.get_total_cost()
        )

    finance_preview.short_description = 'Финансы'

    def changelist_view(self, request, extra_context=None):
        if request.GET.get('view') == 'calendar':
            lessons = self.get_queryset(request).select_related(
                'teacher__user', 'subject'
            ).prefetch_related('enrollments__student__user')

            extra_context = extra_context or {}
            extra_context['lessons'] = lessons
            extra_context['title'] = 'Календарь групповых занятий'

            return render(request, 'admin/school/grouplesson/change_list_calendar.html', extra_context)

        return super().changelist_view(request, extra_context)


# ==================== GROUP ENROLLMENT ADMIN ====================

@admin.register(GroupEnrollment)
class GroupEnrollmentAdmin(admin.ModelAdmin):
    list_display = ('id', 'group_lesson', 'student', 'cost_to_pay', 'status_badge')
    list_filter = ('status', 'group_lesson__subject')
    search_fields = ('student__user__last_name', 'group_lesson__subject__name')
    raw_id_fields = ['student', 'group_lesson']

    def status_badge(self, obj):
        status_colors = {
            'registered': ('#007bff', 'Зарегистрирован'),
            'attended': ('#28a745', 'Присутствовал'),
            'absent': ('#dc3545', 'Отсутствовал'),
            'debt': ('#fd7e14', 'Задолженность'),
        }
        color, text = status_colors.get(obj.status, ('#6c757d', obj.status))
        return format_html(
            '<span style="background: {}; color: white; padding: 3px 8px; border-radius: 3px;">{}</span>',
            color, text)

    status_badge.short_description = 'Статус'


# ==================== ATTENDANCE ADMIN ====================

@admin.register(LessonAttendance)
class LessonAttendanceAdmin(admin.ModelAdmin):
    list_display = ('id', 'lesson_link', 'student', 'cost', 'status_badge', 'registered_at')
    list_filter = ('status', 'lesson__subject')
    search_fields = ('student__user__last_name', 'lesson__subject__name')
    raw_id_fields = ['lesson', 'student']

    def lesson_link(self, obj):
        url = f'/admin/school/lesson/{obj.lesson.id}/change/'
        return format_html('<a href="{}">{} #{}</a>', url, obj.lesson.subject, obj.lesson.id)

    lesson_link.short_description = 'Урок'

    def status_badge(self, obj):
        status_colors = {
            'registered': ('#007bff', 'Зарегистрирован'),
            'attended': ('#28a745', 'Присутствовал'),
            'absent': ('#dc3545', 'Отсутствовал'),
            'debt': ('#fd7e14', 'Задолженность'),
        }
        color, text = status_colors.get(obj.status, ('#6c757d', obj.status))
        return format_html(
            '<span style="background: {}; color: white; padding: 3px 8px; border-radius: 3px;">{}</span>',
            color, text)

    status_badge.short_description = 'Статус'


# ==================== SCHEDULE TEMPLATE ADMIN ====================

@admin.register(ScheduleTemplate)
class ScheduleTemplateAdmin(admin.ModelAdmin):
    list_display = ('id', 'teacher', 'subject', 'start_time', 'repeat_type',
                    'get_days', 'start_date', 'is_active')
    list_filter = ('repeat_type', 'is_active', 'teacher', 'subject')
    search_fields = ('teacher__user__last_name', 'subject__name')
    inlines = [ScheduleTemplateStudentInline]

    fieldsets = (
        ('Основное', {
            'fields': ('teacher', 'subject', 'format', 'start_time', 'end_time')
        }),
        ('Расписание', {
            'fields': ('repeat_type', 'start_date', 'end_date', 'max_occurrences')
        }),
        ('Дни недели', {
            'fields': ('monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'),
            'classes': ('wide',)
        }),
        ('Финансы', {
            'fields': ('base_cost', 'base_teacher_payment'),
            'description': 'Базовая стоимость урока. Индивидуальные цены можно настроить для каждого ученика ниже.'
        }),
        ('Платформа', {
            'fields': ('meeting_link', 'meeting_platform')
        }),
        ('Заметки', {
            'fields': ('notes',),
            'classes': ('collapse',),
        }),
    )

    def get_days(self, obj):
        days = []
        if obj.monday: days.append('Пн')
        if obj.tuesday: days.append('Вт')
        if obj.wednesday: days.append('Ср')
        if obj.thursday: days.append('Чт')
        if obj.friday: days.append('Пт')
        if obj.saturday: days.append('Сб')
        if obj.sunday: days.append('Вс')
        return ', '.join(days) if days else 'Все'

    get_days.short_description = 'Дни'

    actions = ['generate_lessons', 'duplicate_template']

    def generate_lessons(self, request, queryset):
        count = 0
        for template in queryset:
            lessons = template.generate_lessons()
            count += len(lessons)
        self.message_user(request, f'✅ Создано {count} уроков')

    generate_lessons.short_description = '📅 Создать уроки по шаблону'

    def duplicate_template(self, request, queryset):
        for template in queryset:
            template.pk = None
            template.id = None
            template.is_active = True
            template.save()
        self.message_user(request, f'✅ Скопировано {queryset.count()} шаблонов')

    duplicate_template.short_description = '📋 Дублировать шаблон'


# ==================== STUDENT SUBJECT PRICE ADMIN ====================

@admin.register(StudentSubjectPrice)
class StudentSubjectPriceAdmin(admin.ModelAdmin):
    list_display = ['student', 'teacher', 'subject', 'cost', 'teacher_payment', 'discount', 'is_active']
    list_filter = ['subject','teacher', 'is_active']
    search_fields = ['student__user__last_name', 'subject__name']
    list_editable = ['cost', 'teacher_payment', 'is_active']
    autocomplete_fields = ['student','teacher', 'subject']


# =================ЛОГИРОВАНИЕ=============================

@admin.register(UserActionLog)
class UserActionLogAdmin(admin.ModelAdmin):
    list_display = (
    'id', 'created_at_colored', 'user_link', 'action_colored', 'description_short', 'ip_address', 'object_link')
    list_filter = ('action_type', 'created_at', 'user')
    search_fields = ('user__username', 'user__email', 'description', 'ip_address')
    date_hierarchy = 'created_at'
    readonly_fields = ('created_at', 'ip_address', 'user_agent', 'url', 'additional_data')
    list_per_page = 50

    fieldsets = (
        ('Основное', {
            'fields': ('user', 'action_type', 'description', 'created_at')
        }),
        ('Техническая информация', {
            'fields': ('ip_address', 'user_agent', 'url'),
            'classes': ('wide',),
        }),
        ('Связанные объекты', {
            'fields': ('object_type', 'object_id'),
            'classes': ('wide',),
        }),
        ('Дополнительные данные', {
            'fields': ('additional_data',),
            'classes': ('wide', 'collapse'),
        }),
    )

    def created_at_colored(self, obj):
        from django.utils import timezone
        if (timezone.now() - obj.created_at).seconds < 3600:
            color = '#28a745'  # зеленый - свежие
        elif (timezone.now() - obj.created_at).days < 1:
            color = '#ffc107'  # желтый - сегодня
        else:
            color = '#6c757d'  # серый - старые

        return format_html(
            '<span style="color: {};">{}</span>',
            color, obj.created_at.strftime('%d.%m.%Y %H:%M')
        )

    created_at_colored.short_description = 'Дата'
    created_at_colored.admin_order_field = 'created_at'

    def user_link(self, obj):
        url = f'/admin/school/user/{obj.user.id}/change/'
        return format_html('<a href="{}">{}</a>', url, obj.user.get_full_name())

    user_link.short_description = 'Пользователь'

    def action_colored(self, obj):
        colors = {
            'login': '#28a745',
            'logout': '#dc3545',
            'calendar_export': '#17a2b8',
            'lesson_view': '#007bff',
            'video_room_enter': '#ffc107',
        }
        color = colors.get(obj.action_type, '#6c757d')
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            color, obj.get_action_type_display()
        )

    action_colored.short_description = 'Действие'

    def description_short(self, obj):
        return obj.description[:50] + '...' if len(obj.description) > 50 else obj.description

    description_short.short_description = 'Описание'

    def object_link(self, obj):
        if obj.object_type == 'lesson' and obj.object_id:
            url = f'/admin/school/lesson/{obj.object_id}/change/'
            return format_html('<a href="{}">Урок #{}</a>', url, obj.object_id)
        elif obj.object_type == 'homework' and obj.object_id:
            url = f'/admin/school/homework/{obj.object_id}/change/'
            return format_html('<a href="{}">ДЗ #{}</a>', url, obj.object_id)
        return '-'

    object_link.short_description = 'Объект'

    actions = ['export_logs_excel']

    def export_logs_excel(self, request, queryset):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        from datetime import datetime

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Логи действий"

        headers = ['ID', 'Дата', 'Пользователь', 'Действие', 'Описание', 'IP', 'URL']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row, log in enumerate(queryset, start=2):
            ws.cell(row=row, column=1, value=log.id)
            ws.cell(row=row, column=2, value=log.created_at.strftime('%d.%m.%Y %H:%M'))
            ws.cell(row=row, column=3, value=log.user.get_full_name())
            ws.cell(row=row, column=4, value=log.get_action_type_display())
            ws.cell(row=row, column=5, value=log.description)
            ws.cell(row=row, column=6, value=log.ip_address)
            ws.cell(row=row, column=7, value=log.url)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"logs_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    export_logs_excel.short_description = "📥 Экспорт логов в Excel"


@admin.register(PaymentRequest)
class PaymentRequestAdmin(admin.ModelAdmin):
    """
    Административный интерфейс для запросов на выплаты учителям.

    Реализует полный цикл обработки запросов:
        1. Учитель создает запрос (через API или другой интерфейс)
        2. Админ видит запрос в статусе "pending"
        3. Админ может одобрить, отклонить или сразу создать выплату

    Жизненный цикл запроса:
        pending (⏳) → approved (✅) → paid (💰)
                   ↘ rejected (❌)

    Методы:
        approve_request() - одобрение запроса (без выплаты)
        reject_request() - отклонение с указанием причины
        create_payment() - создание фактического платежа и перевод в paid

    Кастомные действия:
        approve_requests() - массовое одобрение
        reject_requests() - массовое отклонение
        mark_as_paid() - массовое создание выплат

    Кнопки действий в списке:
        * Для pending: [Одобрить] [Отклонить]
        * Для approved: [Создать выплату]
    """
    list_display = (
    'id', 'teacher_link', 'amount_colored', 'status_badge', 'payment_method', 'created_at', 'action_buttons')
    list_filter = ('status', 'payment_method', 'created_at')
    search_fields = ('teacher__user__last_name', 'teacher__user__email', 'payment_details')
    readonly_fields = ('created_at', 'updated_at')
    actions = ['approve_requests', 'reject_requests', 'mark_as_paid']

    fieldsets = (
        ('Информация о запросе', {
            'fields': ('teacher', 'amount', 'payment_method', 'payment_details', 'status')
        }),
        ('Обработка', {
            'fields': ('comment', 'payment'),
            'classes': ('wide',),
        }),
        ('Даты', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',),
        }),
    )

    def teacher_link(self, obj):
        url = f'/admin/school/teacher/{obj.teacher.id}/change/'
        return format_html('<a href="{}">{}</a>', url, obj.teacher.user.get_full_name())

    teacher_link.short_description = 'Учитель'

    def amount_colored(self, obj):
        return format_html('<span style="color: #28a745; font-weight: bold;">{} ₽</span>', obj.amount)

    amount_colored.short_description = 'Сумма'

    def status_badge(self, obj):
        colors = {
            'pending': ('#ffc107', '⏳ Ожидает'),
            'approved': ('#17a2b8', '✅ Одобрено'),
            'rejected': ('#dc3545', '❌ Отклонено'),
            'paid': ('#28a745', '💰 Выплачено'),
        }
        color, text = colors.get(obj.status, ('#6c757d', obj.status))
        return format_html(
            '<span style="background: {}; color: white; padding: 3px 8px; border-radius: 3px;">{}</span>',
            color, text
        )

    status_badge.short_description = 'Статус'

    def action_buttons(self, obj):
        if obj.status == 'pending':
            return format_html(
                '<a class="button" href="{}" style="background: #28a745; color: white; padding: 5px 10px; border-radius: 3px; text-decoration: none; margin-right: 5px;">✓ Одобрить</a>'
                '<a class="button" href="{}" style="background: #dc3545; color: white; padding: 5px 10px; border-radius: 3px; text-decoration: none;">✗ Отклонить</a>',
                f'/admin/school/paymentrequest/{obj.id}/approve/',
                f'/admin/school/paymentrequest/{obj.id}/reject/'
            )
        elif obj.status == 'approved':
            return format_html(
                '<a class="button" href="{}" style="background: #17a2b8; color: white; padding: 5px 10px; border-radius: 3px; text-decoration: none;">💰 Создать выплату</a>',
                f'/admin/school/paymentrequest/{obj.id}/create-payment/'
            )
        return '-'

    action_buttons.short_description = 'Действия'

    def get_urls(self):
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('<int:request_id>/approve/', self.admin_site.admin_view(self.approve_request),
                 name='paymentrequest-approve'),
            path('<int:request_id>/reject/', self.admin_site.admin_view(self.reject_request),
                 name='paymentrequest-reject'),
            path('<int:request_id>/create-payment/', self.admin_site.admin_view(self.create_payment),
                 name='paymentrequest-create-payment'),
        ]
        return custom_urls + urls

    def approve_request(self, request, request_id):
        from django.shortcuts import get_object_or_404, redirect
        from django.contrib import messages

        payment_request = get_object_or_404(PaymentRequest, id=request_id)
        payment_request.approve(request.user)
        messages.success(request, f'Запрос #{request_id} одобрен')
        return redirect('admin:school_paymentrequest_changelist')

    def reject_request(self, request, request_id):
        from django.shortcuts import get_object_or_404, redirect, render
        from django.contrib import messages

        payment_request = get_object_or_404(PaymentRequest, id=request_id)

        if request.method == 'POST':
            reason = request.POST.get('reason', '')
            payment_request.reject(request.user, reason)
            messages.success(request, f'Запрос #{request_id} отклонен')
            return redirect('admin:school_paymentrequest_changelist')

        context = {
            'title': f'Отклонить запрос #{request_id}',
            'payment_request': payment_request,
        }
        return render(request, 'admin/school/paymentrequest/reject_form.html', context)

    def create_payment(self, request, request_id):
        from django.shortcuts import get_object_or_404, redirect, render
        from django.contrib import messages
        from decimal import Decimal

        payment_request = get_object_or_404(PaymentRequest, id=request_id)

        if request.method == 'POST':
            # Создаем платеж
            payment = Payment.objects.create(
                user=payment_request.teacher.user,
                amount=payment_request.amount,
                payment_type='teacher_payment',
                description=f'Выплата по запросу #{payment_request.id} ({payment_request.payment_method})'
            )

            # Уменьшаем баланс учителя
            teacher = payment_request.teacher
            teacher.wallet_balance -= payment_request.amount
            teacher.save()

            # Отмечаем запрос как выполненный
            payment_request.mark_as_paid(request.user, payment)

            messages.success(request, f'✅ Выплата #{payment.id} создана, баланс учителя уменьшен')
            return redirect('admin:school_paymentrequest_changelist')

        context = {
            'title': f'Создать выплату по запросу #{request_id}',
            'payment_request': payment_request,
            'teacher_balance': payment_request.teacher.wallet_balance,
        }
        return render(request, 'admin/school/paymentrequest/create_payment.html', context)

    def approve_requests(self, request, queryset):
        for req in queryset.filter(status='pending'):
            req.approve(request.user)
        self.message_user(request, f'✅ Одобрено {queryset.count()} запросов')

    approve_requests.short_description = "✅ Одобрить выбранные запросы"

    def reject_requests(self, request, queryset):
        for req in queryset.filter(status='pending'):
            req.reject(request.user, 'Отклонено администратором')
        self.message_user(request, f'❌ Отклонено {queryset.count()} запросов')

    reject_requests.short_description = "❌ Отклонить выбранные запросы"

    def mark_as_paid(self, request, queryset):
        count = 0
        for req in queryset.filter(status='approved'):
            # Создаем платеж
            payment = Payment.objects.create(
                user=req.teacher.user,
                amount=req.amount,
                payment_type='teacher_payment',
                description=f'Выплата по запросу #{req.id}'
            )
            req.mark_as_paid(request.user, payment)
            count += 1
        self.message_user(request, f'💰 Создано {count} выплат')

    mark_as_paid.short_description = "💰 Создать выплаты по одобренным запросам"


@admin.register(Feedback)
class FeedbackAdmin(admin.ModelAdmin):
    list_display = ('name', 'role', 'rating_display', 'is_active', 'is_on_main', 'sort_order', 'created_at')
    list_filter = ('is_active', 'is_on_main', 'rating', 'teacher')
    search_fields = ('name', 'text', 'role')
    list_editable = ('is_active', 'is_on_main', 'sort_order')
    list_per_page = 20

    fieldsets = (
        ('Информация об отзыве', {
            'fields': ('name', 'role', 'text', 'photo')
        }),
        ('Оценка и связь', {
            'fields': ('rating', 'teacher'),
            'description': 'Оценка от 1 до 5 звезд. Если отзыв о конкретном учителе - выберите его'
        }),
        ('Настройки отображения', {
            'fields': ('is_active', 'is_on_main', 'sort_order'),
            'description': 'is_active - отображать на сайте, is_on_main - показывать в карусели на главной, sort_order - порядок сортировки (меньше = выше)'
        }),
    )

    def rating_display(self, obj):
        return '★' * obj.rating + '☆' * (5 - obj.rating)

    rating_display.short_description = 'Оценка'
    rating_display.admin_order_field = 'rating'

    actions = ['activate_feedbacks', 'deactivate_feedbacks', 'show_on_main', 'hide_from_main']

    def activate_feedbacks(self, request, queryset):
        queryset.update(is_active=True)
        self.message_user(request, f'✅ {queryset.count()} отзывов активировано')

    activate_feedbacks.short_description = "✅ Активировать выбранные отзывы"

    def deactivate_feedbacks(self, request, queryset):
        queryset.update(is_active=False)
        self.message_user(request, f'❌ {queryset.count()} отзывов деактивировано')

    deactivate_feedbacks.short_description = "❌ Деактивировать выбранные отзывы"

    def show_on_main(self, request, queryset):
        queryset.update(is_on_main=True)
        self.message_user(request, f'🏠 {queryset.count()} отзывов будут показываться на главной')

    show_on_main.short_description = "🏠 Показывать на главной"

    def hide_from_main(self, request, queryset):
        queryset.update(is_on_main=False)
        self.message_user(request, f'🚫 {queryset.count()} отзывов скрыты с главной')

    hide_from_main.short_description = "🚫 Скрыть с главной"


# ==================== REGISTER ALL MODELS ====================
# Регистрация всех моделей в административном интерфейсе
admin.site.register(User, CustomUserAdmin)
admin.site.register(Subject, SubjectAdmin)
admin.site.register(Student, StudentAdmin)
admin.site.register(LessonFormat, LessonFormatAdmin)

# Настройка заголовков админки
admin.site.site_header = 'Плюс Прогресс - Администрирование'
admin.site.site_title = 'Плюс Прогресс'
admin.site.index_title = 'Управление онлайн школой'


