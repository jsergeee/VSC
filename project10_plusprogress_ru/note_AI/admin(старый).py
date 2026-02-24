# school/admin.py

from django.contrib import admin
from django.contrib.auth.admin import UserAdmin
from django.utils.html import format_html
from django.contrib.auth.models import User as AuthUser
from django.urls import path
from django.shortcuts import render, redirect
from django import forms
from django.db import models
from django.utils import timezone
from .models import GroupLesson, GroupEnrollment
from .models import LessonAttendance
from .models import ScheduleTemplate, ScheduleTemplateStudent
from .models import StudentSubjectPrice
from datetime import datetime
from django.db.models import Prefetch
from .models import LessonAttendance

from .models import (
    User, Subject, Teacher, Student, Lesson, LessonFormat,
    LessonReport, Payment, Schedule, TrialRequest,
    Notification, LessonFeedback, TeacherRating,
    Homework, HomeworkSubmission
)
from .views import schedule_calendar_data, admin_complete_lesson

# Разрегистрируем стандартного User, если зарегистрирован
try:
    admin.site.unregister(AuthUser)
except admin.sites.NotRegistered:
    pass


class StudentSubjectPriceInline(admin.TabularInline):
    model = StudentSubjectPrice
    extra = 1
    fields = ['subject', 'cost', 'teacher_payment', 'discount', 'is_active']
    autocomplete_fields = ['subject']


# ==================== CUSTOM USER ADMIN ====================
class CustomUserAdmin(UserAdmin):
    list_display = ('username', 'get_full_name', 'email', 'phone', 'role',
                    'balance', 'is_email_verified_badge', 'is_staff')
    list_filter = ('role', 'is_email_verified', 'is_staff', 'is_superuser', 'groups')
    search_fields = ('username', 'first_name', 'last_name', 'email', 'phone')
    readonly_fields = ('email_verification_sent',)

    fieldsets = (
        (None, {'fields': ('username', 'password')}),
        ('Личная информация', {
            'fields': ('first_name', 'last_name', 'patronymic', 'email', 'phone', 'photo')
        }),
        ('Роль и баланс', {
            'fields': ('role', 'balance'),
            'classes': ('wide',),
        }),
        ('✅ Email подтверждение', {  # ✅ НОВЫЙ БЛОК С ЭМОДЗИ
            'fields': ('is_email_verified', 'email_verification_sent'),
            'classes': ('wide',),
            'description': 'Управление подтверждением email пользователя',
        }),
        ('Права доступа', {
            'fields': ('is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions'),
        }),
        ('Важные даты', {'fields': ('last_login', 'date_joined')}),
    )

    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': ('username', 'password1', 'password2',
                       'first_name', 'last_name', 'patronymic',
                       'email', 'phone', 'role'),
        }),
    )

    # Добавляем действия для массовой обработки
    actions = ['mark_as_verified', 'mark_as_unverified']

    def get_full_name(self, obj):
        full_name = obj.get_full_name()
        if obj.patronymic:
            return f"{full_name} {obj.patronymic}"
        return full_name or obj.username

    get_full_name.short_description = 'ФИО'

    def is_email_verified_badge(self, obj):
        """Отображает статус подтверждения email в виде цветного значка"""
        if obj.is_email_verified:
            return format_html(
                '<span style="background: #28a745; color: white; padding: 3px 8px; border-radius: 3px;">✅ Подтвержден</span>'
            )
        else:
            return format_html(
                '<span style="background: #dc3545; color: white; padding: 3px 8px; border-radius: 3px;">❌ Не подтвержден</span>'
            )

    is_email_verified_badge.short_description = 'Email подтвержден'

    def mark_as_verified(self, request, queryset):
        """Отметить выбранных пользователей как верифицированных"""
        updated = queryset.update(is_email_verified=True)
        self.message_user(request, f'✅ {updated} пользователей отмечены как подтвержденные')

    mark_as_verified.short_description = "✅ Отметить как подтвержденные email"

    def mark_as_unverified(self, request, queryset):
        """Отметить выбранных пользователей как неверифицированных"""
        updated = queryset.update(is_email_verified=False)
        self.message_user(request, f'⚠️ {updated} пользователей отмечены как неподтвержденные')

    mark_as_unverified.short_description = "❌ Отметить как неподтвержденные email"


# ==================== SUBJECT ADMIN ====================
class SubjectAdmin(admin.ModelAdmin):
    list_display = ('name', 'description')
    search_fields = ('name',)


# ==================== TEACHER ADMIN ====================
class TeacherAdmin(admin.ModelAdmin):
    list_display = ('id', 'user', 'display_subjects', 'experience', 'created')
    list_filter = ('subjects',)
    search_fields = ('user__first_name', 'user__last_name', 'user__email')
    filter_horizontal = ('subjects',)

    fieldsets = (
        (None, {
            'fields': ('user', 'subjects', 'experience')
        }),
        ('Дополнительная информация', {
            'fields': ('education', 'bio', 'wallet_balance', 'payment_details'),
            'classes': ('collapse',),
        }),
    )

    def display_subjects(self, obj):
        return ", ".join([s.name for s in obj.subjects.all()])

    display_subjects.short_description = 'Предметы'

    def created(self, obj):
        return obj.user.date_joined.strftime('%d.%m.%Y')

    created.short_description = 'Дата регистрации'

    change_list_template = "admin/school/teacher/change_list.html"
    actions = ['export_teachers_excel']

    def export_teachers_excel(self, request, queryset):
        """Экспорт выбранных учителей в Excel"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Учителя"

        # Заголовки
        headers = ['ID', 'Фамилия', 'Имя', 'Отчество', 'Email', 'Телефон',
                   'Предметы', 'Опыт', 'Баланс кошелька']

        # Стиль заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Данные
        for row, teacher in enumerate(queryset, start=2):
            subjects = ", ".join([s.name for s in teacher.subjects.all()])

            ws.cell(row=row, column=1, value=teacher.id)
            ws.cell(row=row, column=2, value=teacher.user.last_name)
            ws.cell(row=row, column=3, value=teacher.user.first_name)
            ws.cell(row=row, column=4, value=teacher.user.patronymic)
            ws.cell(row=row, column=5, value=teacher.user.email)
            ws.cell(row=row, column=6, value=teacher.user.phone)
            ws.cell(row=row, column=7, value=subjects)
            ws.cell(row=row, column=8, value=teacher.experience)
            ws.cell(row=row, column=9, value=float(teacher.wallet_balance))

        # Настройка ширины колонок
        column_widths = [8, 15, 15, 15, 25, 15, 30, 8, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Создаем response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"teachers_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    export_teachers_excel.short_description = "📥 Экспорт выбранных учителей в Excel"


# ==================== STUDENT ADMIN ====================
class StudentAdmin(admin.ModelAdmin):
    list_display = ('id', 'user', 'parent_name', 'parent_phone', 'get_teachers_count', 'get_balance_display')
    search_fields = ('user__first_name', 'user__last_name', 'user__email', 'parent_name')
    filter_horizontal = ('teachers',)
    list_filter = ('teachers',)
    raw_id_fields = ('user',)
    inlines = [StudentSubjectPriceInline]

    def get_teachers_count(self, obj):
        return obj.teachers.count()

    get_teachers_count.short_description = 'Кол-во учителей'

    def get_balance_display(self, obj):
        balance = obj.user.balance
        if balance > 0:
            return format_html('<span style="color: #28a745;">💰 {}</span>', f"{balance:.2f}")
        elif balance < 0:
            return format_html('<span style="color: #dc3545;">🔴 {}</span>', f"{balance:.2f}")
        return format_html('<span style="color: #6c757d;">⚪ 0.00</span>')

    get_balance_display.short_description = 'Баланс'

    change_list_template = "admin/school/student/change_list.html"
    actions = ['export_students_excel']

    def export_students_excel(self, request, queryset):
        """Экспорт выбранных учеников в Excel"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ученики"

        # Заголовки
        headers = ['ID', 'Фамилия', 'Имя', 'Отчество', 'Email', 'Телефон',
                   'Родитель', 'Телефон родителя', 'Баланс', 'Учителя']

        # Стиль заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Данные
        for row, student in enumerate(queryset, start=2):
            teachers = ", ".join([t.user.get_full_name() for t in student.teachers.all()[:3]])
            if student.teachers.count() > 3:
                teachers += f" и еще {student.teachers.count() - 3}"

            ws.cell(row=row, column=1, value=student.id)
            ws.cell(row=row, column=2, value=student.user.last_name)
            ws.cell(row=row, column=3, value=student.user.first_name)
            ws.cell(row=row, column=4, value=student.user.patronymic)
            ws.cell(row=row, column=5, value=student.user.email)
            ws.cell(row=row, column=6, value=student.user.phone)
            ws.cell(row=row, column=7, value=student.parent_name)
            ws.cell(row=row, column=8, value=student.parent_phone)
            ws.cell(row=row, column=9, value=float(student.user.balance))
            ws.cell(row=row, column=10, value=teachers)

        # Настройка ширины колонок
        column_widths = [8, 15, 15, 15, 25, 15, 20, 15, 12, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Создаем response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"students_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    export_students_excel.short_description = "📥 Экспорт выбранных учеников в Excel"


# ==================== LESSON FORMAT ADMIN ====================
class LessonFormatAdmin(admin.ModelAdmin):
    list_display = ('name', 'description')
    search_fields = ('name',)


# ==================== INLINE ДЛЯ ПОСЕЩАЕМОСТИ ====================
class LessonAttendanceInline(admin.TabularInline):
    model = LessonAttendance
    extra = 1
    raw_id_fields = ['student']
    fields = ['student', 'cost', 'discount', 'teacher_payment_share', 'status']


# ==================== LESSON ADMIN ====================
@admin.register(Lesson)
class LessonAdmin(admin.ModelAdmin):
    formfield_overrides = {
        models.TimeField: {'widget': forms.TimeInput(format='%H:%M', attrs={'type': 'time'})},
        models.DateField: {'widget': forms.DateInput(attrs={'type': 'date'})},
    }

    list_display = ('id', 'subject', 'teacher', 'students_list', 'date', 'start_time', 'status', 'get_total_cost')
    list_filter = ('status', 'subject', 'date', 'teacher', 'is_group')
    search_fields = ('teacher__user__last_name', 'students__user__last_name', 'subject__name')
    date_hierarchy = 'date'
    raw_id_fields = ('teacher',)
    inlines = [LessonAttendanceInline]

    fieldsets = (
        ('Основная информация', {
            'fields': ('teacher', 'subject', 'format', 'is_group')
        }),
        ('Время', {
            'fields': ('date', 'start_time', 'end_time')
        }),
        ('Финансы', {
            'fields': ('price_type', 'base_cost', 'base_teacher_payment')
        }),
        ('Платформа', {
            'fields': ('meeting_link', 'meeting_platform', 'video_room')
        }),
        ('Статус', {
            'fields': ('status', 'notes')
        }),
    )

    def students_list(self, obj):
        """Отображает список учеников"""
        students = obj.students.all()
        if not students:
            return "—"
        elif students.count() == 1:
            return students.first().user.get_full_name()
        else:
            return f"{students.count()} учеников"

    students_list.short_description = 'Ученики'

    def get_total_cost(self, obj):
        """Общая стоимость урока"""
        return obj.get_total_cost()

    get_total_cost.short_description = 'Общая стоимость'

    def has_report(self, obj):
        if hasattr(obj, 'report'):
            url = f'/admin/school/lessonreport/{obj.report.id}/change/'
            return format_html('<a href="{}" style="color: #28a745;">✅ Отчет #{}</a>', url, obj.report.id)
        return '❌ Нет отчета'

    has_report.short_description = 'Отчет'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('<int:lesson_id>/complete/',
                 self.admin_site.admin_view(admin_complete_lesson),
                 name='complete-lesson'),
        ]
        return custom_urls + urls

    from django.db.models import Prefetch
    from .models import LessonAttendance

    def changelist_view(self, request, extra_context=None):
        # Если запрошен календарь
        if request.GET.get('view') == 'calendar':
            lessons = self.get_queryset(request).select_related(
                'teacher__user', 'subject'
            ).prefetch_related(
                Prefetch(
                    'attendance',
                    queryset=LessonAttendance.objects.select_related('student__user')
                )
            )

            # Формируем события для календаря
            calendar_events = []
            for lesson in lessons:
                # Предмет (сокращенно - первые 2 буквы)
                subject_short = lesson.subject.name[:4]

                # Фамилия учителя
                teacher_last = lesson.teacher.user.last_name

                # Ученики
                attendance_count = lesson.attendance.count()
                if attendance_count == 0:
                    students_text = "нет"
                elif attendance_count == 1:
                    student = lesson.attendance.first().student
                    first_name = student.user.first_name
                    last_initial = student.user.last_name[0] if student.user.last_name else ''
                    students_text = f"{first_name} {last_initial}."
                else:
                    students_text = f"{attendance_count} уч."

                # Формируем заголовок
                title = f"{subject_short} {teacher_last} - {students_text}"

                # Определяем цвет в зависимости от статуса
                if lesson.status == 'completed':
                    bg_color = '#28a745'
                elif lesson.status == 'cancelled':
                    bg_color = '#dc3545'
                elif lesson.status == 'overdue':
                    bg_color = '#fd7e14'
                else:
                    bg_color = '#007bff'

                calendar_events.append({
                    'title': title,
                    'start': f"{lesson.date}T{lesson.start_time}",
                    'end': f"{lesson.date}T{lesson.end_time}",
                    'url': f"/admin/school/lesson/{lesson.id}/change/",
                    'backgroundColor': bg_color,
                    'borderColor': bg_color,
                    'textColor': 'white',
                })

            extra_context = extra_context or {}
            extra_context['calendar_events'] = calendar_events
            extra_context['title'] = 'Календарь занятий'

            return render(request, 'admin/school/lesson/change_list_calendar.html', extra_context)

        # Обычный список
        return super().changelist_view(request, extra_context)

    change_form_template = "admin/school/lesson/change_form.html"

    def response_change(self, request, obj):
        if "_complete-lesson" in request.POST:
            # Перенаправляем на страницу завершения
            return redirect('admin-complete-lesson', lesson_id=obj.id)
        return super().response_change(request, obj)



# ==================== LESSON REPORT ADMIN ====================
@admin.register(LessonReport)
class LessonReportAdmin(admin.ModelAdmin):
    list_display = ('id', 'lesson_link', 'topic_preview', 'created_at')
    list_filter = ('created_at', 'lesson__subject')
    search_fields = ('topic', 'lesson__subject__name')
    readonly_fields = ('created_at',)
    raw_id_fields = ('lesson',)

    def lesson_link(self, obj):
        url = f'/admin/school/lesson/{obj.lesson.id}/change/'
        return format_html('<a href="{}">{} #{}</a>', url, obj.lesson.subject, obj.lesson.id)

    lesson_link.short_description = 'Занятие'

    def topic_preview(self, obj):
        return obj.topic[:50] + '...' if len(obj.topic) > 50 else obj.topic

    topic_preview.short_description = 'Тема'


# ==================== PAYMENT ADMIN ====================
@admin.register(Payment)
class PaymentAdmin(admin.ModelAdmin):
    list_display = ('user', 'amount', 'payment_type', 'description', 'created_at')
    list_filter = ('payment_type', 'created_at')
    search_fields = ('user__first_name', 'user__last_name', 'description')
    date_hierarchy = 'created_at'
    raw_id_fields = ('user', 'lesson')

    def save_model(self, request, obj, form, change):
        # Сохраняем платеж
        super().save_model(request, obj, form, change)

        # Обновляем баланс пользователя
        if obj.payment_type == 'income':
            obj.user.balance += obj.amount
        elif obj.payment_type == 'expense':
            obj.user.balance -= obj.amount

        obj.user.save()

        # Создаем уведомление
        from .models import Notification

        if obj.payment_type == 'income':
            title = '💰 Пополнение баланса'
            message = f'Ваш баланс пополнен на {obj.amount} ₽'
        elif obj.payment_type == 'expense':
            title = '💸 Списание средств'
            message = f'С вашего баланса списано {obj.amount} ₽'
        else:
            title = '💳 Выплата'
            message = f'Вам начислена выплата {obj.amount} ₽'

        Notification.objects.create(
            user=obj.user,
            title=title,
            message=message,
            notification_type='payment_received' if obj.payment_type == 'income' else 'payment_withdrawn',
            link='/profile/'
        )


# ==================== SCHEDULE ADMIN ====================
@admin.register(Schedule)
class ScheduleAdmin(admin.ModelAdmin):
    formfield_overrides = {
        models.TimeField: {'widget': forms.TimeInput(format='%H:%M', attrs={'type': 'time'})},
        models.DateField: {'widget': forms.DateInput(attrs={'type': 'date'})},
    }

    list_display = ('teacher', 'date', 'day_of_week_display', 'start_time', 'end_time', 'is_active')
    list_filter = ('date', 'is_active', 'teacher')
    search_fields = ('teacher__user__last_name',)

    def day_of_week_display(self, obj):
        days = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
        return days[obj.date.weekday()]

    day_of_week_display.short_description = 'День'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('calendar-data/', self.admin_site.admin_view(schedule_calendar_data), name='schedule-calendar-data'),
        ]
        return custom_urls + urls


# ==================== TRIAL REQUEST ADMIN ====================
@admin.register(TrialRequest)
class TrialRequestAdmin(admin.ModelAdmin):
    list_display = ('name', 'phone', 'email', 'subject', 'created_at', 'is_processed')
    list_filter = ('is_processed', 'subject', 'created_at')
    search_fields = ('name', 'phone', 'email')
    date_hierarchy = 'created_at'


# ==================== NOTIFICATION ADMIN ====================
@admin.register(Notification)
class NotificationAdmin(admin.ModelAdmin):
    list_display = ('id', 'user', 'title', 'notification_type', 'is_read', 'created_at')
    list_filter = ('notification_type', 'is_read', 'created_at')
    search_fields = ('user__username', 'user__email', 'title')
    date_hierarchy = 'created_at'
    raw_id_fields = ('user',)
    list_per_page = 50


# ==================== LESSON FEEDBACK ADMIN ====================
@admin.register(LessonFeedback)
class LessonFeedbackAdmin(admin.ModelAdmin):
    list_display = ('id', 'lesson', 'student', 'teacher', 'rating_stars', 'created_at', 'is_public')
    list_filter = ('rating', 'is_public', 'created_at')
    search_fields = ('student__user__last_name', 'teacher__user__last_name', 'comment')
    raw_id_fields = ('lesson', 'student', 'teacher')
    date_hierarchy = 'created_at'
    actions = ['make_public', 'make_private']

    def rating_stars(self, obj):
        return '⭐' * obj.rating

    rating_stars.short_description = 'Оценка'

    def make_public(self, request, queryset):
        queryset.update(is_public=True)
        self.message_user(request, f'✅ {queryset.count()} оценок опубликовано')

    make_public.short_description = 'Опубликовать'

    def make_private(self, request, queryset):
        queryset.update(is_public=False)
        self.message_user(request, f'🔒 {queryset.count()} оценок скрыто')

    make_private.short_description = 'Скрыть'


# ==================== TEACHER RATING ADMIN ====================
@admin.register(TeacherRating)
class TeacherRatingAdmin(admin.ModelAdmin):
    list_display = ('teacher', 'average_rating_display', 'total_feedbacks', 'updated_at')
    list_select_related = ('teacher__user',)
    readonly_fields = ('teacher', 'average_rating', 'total_feedbacks', 'rating_5_count',
                       'rating_4_count', 'rating_3_count', 'rating_2_count', 'rating_1_count', 'updated_at')

    def average_rating_display(self, obj):
        return f"{obj.average_rating:.1f} ⭐"

    average_rating_display.short_description = 'Средний балл'


# ==================== HOMEWORK ADMIN ====================
@admin.register(Homework)
class HomeworkAdmin(admin.ModelAdmin):
    list_display = ('id', 'colored_title', 'student_name', 'teacher_name', 'subject',
                    'colored_deadline', 'colored_status')
    list_filter = ('subject', 'is_active', 'deadline')
    search_fields = ('title', 'student__user__last_name', 'teacher__user__last_name')
    date_hierarchy = 'deadline'
    raw_id_fields = ('student', 'teacher', 'subject')
    list_per_page = 25
    save_on_top = True

    def student_name(self, obj):
        return obj.student.user.get_full_name()

    student_name.short_description = 'Ученик'

    def teacher_name(self, obj):
        return obj.teacher.user.get_full_name()

    teacher_name.short_description = 'Учитель'

    def colored_title(self, obj):
        return format_html('<span style="color: #2c3e50; font-weight: bold;">{}</span>', obj.title)

    colored_title.short_description = 'Название'

    def colored_deadline(self, obj):
        now = timezone.now()
        if obj.deadline < now:
            return format_html('<span style="color: #dc3545;">⚠️ {}</span>',
                               obj.deadline.strftime('%d.%m.%Y %H:%M'))
        elif (obj.deadline - now).days < 1:
            return format_html('<span style="color: #ffc107;">⚡ {}</span>',
                               obj.deadline.strftime('%d.%m.%Y %H:%M'))
        else:
            return format_html('<span style="color: #28a745;">✅ {}</span>',
                               obj.deadline.strftime('%d.%m.%Y %H:%M'))

    colored_deadline.short_description = 'Срок сдачи'

    def colored_status(self, obj):
        status = obj.get_status()
        colors = {
            'pending': ('#ffc107', '⏳ Ожидает'),
            'submitted': ('#17a2b8', '📤 На проверке'),
            'checked': ('#28a745', '✅ Проверено'),
            'overdue': ('#dc3545', '⚠️ Просрочено'),
        }
        color, text = colors.get(status, ('#6c757d', '❓'))
        return format_html(
            '<span style="background: {}; color: white; padding: 3px 8px; border-radius: 3px;">{}</span>',
            color, text)

    colored_status.short_description = 'Статус'


# ==================== HOMEWORK SUBMISSION ADMIN ====================
@admin.register(HomeworkSubmission)
class HomeworkSubmissionAdmin(admin.ModelAdmin):
    list_display = ('id', 'homework_link', 'student_name', 'submitted_at', 'status_colored', 'grade_display')
    list_filter = ('status', 'submitted_at')
    search_fields = ('homework__title', 'student__user__last_name')
    date_hierarchy = 'submitted_at'

    def homework_link(self, obj):
        url = f'/admin/school/homework/{obj.homework.id}/change/'
        return format_html('<a href="{}">{}</a>', url, obj.homework.title)

    homework_link.short_description = 'Задание'

    def student_name(self, obj):
        return obj.student.user.get_full_name()

    student_name.short_description = 'Ученик'

    def status_colored(self, obj):
        if obj.status == 'submitted':
            return format_html('<span style="color: #17a2b8;">📤 Ожидает проверки</span>')
        return format_html('<span style="color: #28a745;">✅ Проверено</span>')

    status_colored.short_description = 'Статус'

    def grade_display(self, obj):
        if obj.grade:
            return format_html(
                '<span style="background: #28a745; color: white; padding: 2px 6px; border-radius: 3px;">{}/5</span>',
                obj.grade)
        return '—'

    grade_display.short_description = 'Оценка'


# ==================== GROUP LESSON ADMIN ====================
class GroupEnrollmentInline(admin.TabularInline):
    model = GroupEnrollment
    extra = 1
    raw_id_fields = ['student']


@admin.register(GroupLesson)
class GroupLessonAdmin(admin.ModelAdmin):
    list_display = ('id', 'subject', 'teacher', 'date', 'start_time', 'students_count', 'status', 'get_total_cost')
    list_filter = ('status', 'subject', 'teacher', 'date')
    search_fields = ('subject__name', 'teacher__user__last_name', 'notes')
    inlines = [GroupEnrollmentInline]
    fieldsets = (
        ('Основное', {
            'fields': ('teacher', 'subject', 'format', 'date', 'start_time', 'end_time')
        }),
        ('Финансы', {
            'fields': ('price_type', 'base_price', 'teacher_payment')
        }),
        ('Платформа', {
            'fields': ('meeting_link', 'meeting_platform', 'video_room')
        }),
        ('Статус', {
            'fields': ('status', 'notes')
        }),
    )

    def students_count(self, obj):
        return obj.enrollments.count()

    students_count.short_description = 'Учеников'

    def get_total_cost(self, obj):
        return obj.get_total_cost()

    get_total_cost.short_description = 'Общая стоимость'

    def changelist_view(self, request, extra_context=None):
        if request.GET.get('view') == 'calendar':
            lessons = self.get_queryset(request).select_related(
                'teacher__user', 'subject'
            ).prefetch_related('enrollments__student__user')

            extra_context = extra_context or {}
            extra_context['lessons'] = lessons
            extra_context['title'] = 'Календарь групповых занятий'

            return render(request, 'admin/school/grouplesson/change_list_calendar.html', extra_context)

        return super().changelist_view(request, extra_context)


# ==================== GROUP ENROLLMENT ADMIN ====================
@admin.register(GroupEnrollment)
class GroupEnrollmentAdmin(admin.ModelAdmin):
    list_display = ('id', 'group_lesson', 'student', 'cost_to_pay', 'status')
    list_filter = ('status', 'group_lesson__subject')
    search_fields = ('student__user__last_name', 'group_lesson__subject__name')
    raw_id_fields = ['student', 'group_lesson']


# ==================== ATTENDANCE ADMIN ====================
@admin.register(LessonAttendance)
class LessonAttendanceAdmin(admin.ModelAdmin):
    list_display = ('id', 'lesson', 'student', 'cost', 'status')
    list_filter = ('status', 'lesson__subject')
    search_fields = ('student__user__last_name', 'lesson__subject__name')
    raw_id_fields = ['lesson', 'student']


# ==================== SCHEDULE TEMPLATE ADMIN ====================
class ScheduleTemplateStudentInline(admin.TabularInline):
    model = ScheduleTemplateStudent
    extra = 1
    raw_id_fields = ['student']


@admin.register(ScheduleTemplate)
class ScheduleTemplateAdmin(admin.ModelAdmin):
    list_display = ('id', 'teacher', 'subject', 'start_time', 'repeat_type', 'get_days', 'start_date', 'is_active')
    list_filter = ('repeat_type', 'is_active', 'teacher', 'subject')
    search_fields = ('teacher__user__last_name', 'subject__name')
    inlines = [ScheduleTemplateStudentInline]
    fieldsets = (
        ('Основное', {
            'fields': ('teacher', 'subject', 'format', 'start_time', 'end_time')
        }),
        ('Расписание', {
            'fields': ('repeat_type', 'start_date', 'end_date', 'max_occurrences')
        }),
        ('Дни недели', {
            'fields': ('monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'),
            'classes': ('wide',)
        }),
        ('Финансы', {
            'fields': ('price_type', 'base_cost', 'base_teacher_payment')
        }),
        ('Платформа', {
            'fields': ('meeting_link', 'meeting_platform')
        }),
    )

    def get_days(self, obj):
        days = []
        if obj.monday: days.append('Пн')
        if obj.tuesday: days.append('Вт')
        if obj.wednesday: days.append('Ср')
        if obj.thursday: days.append('Чт')
        if obj.friday: days.append('Пт')
        if obj.saturday: days.append('Сб')
        if obj.sunday: days.append('Вс')
        return ', '.join(days) if days else 'Все'

    get_days.short_description = 'Дни'

    actions = ['generate_lessons']

    def generate_lessons(self, request, queryset):
        count = 0
        for template in queryset:
            lessons = template.generate_lessons()
            count += len(lessons)
        self.message_user(request, f'Создано {count} уроков')

    generate_lessons.short_description = 'Создать уроки по шаблону'


# ==================== STUDENT SUBJECT PRICE ADMIN ====================
@admin.register(StudentSubjectPrice)
class StudentSubjectPriceAdmin(admin.ModelAdmin):
    list_display = ['student', 'subject', 'cost', 'teacher_payment', 'discount', 'is_active']
    list_filter = ['subject', 'is_active']
    search_fields = ['student__user__last_name', 'student__user__first_name', 'subject__name']
    list_editable = ['cost', 'teacher_payment', 'is_active']
    autocomplete_fields = ['student', 'subject']
    date_hierarchy = 'created_at'


# ==================== REGISTER ALL MODELS ====================
admin.site.register(User, CustomUserAdmin)
admin.site.register(Subject, SubjectAdmin)
admin.site.register(Teacher, TeacherAdmin)
admin.site.register(Student, StudentAdmin)
admin.site.register(LessonFormat, LessonFormatAdmin)
# Lesson уже зарегистрирован через @admin.register(Lesson)
# LessonReport уже зарегистрирован через @admin.register(LessonReport)
# Payment уже зарегистрирован через @admin.register(Payment)
# Schedule уже зарегистрирован через @admin.register(Schedule)
# TrialRequest уже зарегистрирован через @admin.register(TrialRequest)
# Notification уже зарегистрирован через @admin.register(Notification)
# LessonFeedback уже зарегистрирован через @admin.register(LessonFeedback)
# TeacherRating уже зарегистрирован через @admin.register(TeacherRating)
# Homework уже зарегистрирован через @admin.register(Homework)
# HomeworkSubmission уже зарегистрирован через @admin.register(HomeworkSubmission)
# GroupLesson уже зарегистрирован через @admin.register(GroupLesson)
# GroupEnrollment уже зарегистрирован через @admin.register(GroupEnrollment)
# LessonAttendance уже зарегистрирован через @admin.register(LessonAttendance)
# ScheduleTemplate уже зарегистрирован через @admin.register(ScheduleTemplate)
# StudentSubjectPrice уже зарегистрирован через @admin.register(StudentSubjectPrice)

# Настройка заголовков админки
admin.site.site_header = 'Плюс Прогресс - Администрирование'
admin.site.site_title = 'Плюс Прогресс'
admin.site.index_title = 'Управление онлайн школой'
